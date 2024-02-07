from blink import *
import re, os, sys, itertools, string, time
import numpy as np
import pandas as pd
import tempfile
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtWidgets import QFileDialog, QLabel,QMessageBox, QWidget
import matplotlib.pyplot as plt
from matplotlib.pyplot import axes, figure
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

from HealthSubCentreVal import *
# from PrimaryHealthCentre import *
# from SubDistrictHospitalVal import *
# from DistrictHospitalVal import *
# from CommunityHealthCentreVal import *

from After_Upload_New import *
from Downloader import *
from Validater import *

from pandas.io.formats import style
import openpyxl
from openpyxl import load_workbook
from PyQt5.QtWidgets import QWidget, QProgressBar, QVBoxLayout
import tempfile
from collections import Counter

#!###########################################################
### SQL Imports

#!###########################################################


# # ## PROGRESS BAR
# # # ============
# class PopUpProgressBar(QtWidgets.QWidget):

#     def __init__(self):
#         super().__init__()

#         layout = QtWidgets.QVBoxLayout()
        
#         self.widget = QtWidgets.QWidget(self)
#         layout.addWidget(self.widget)

#         self.label = QtWidgets.QLabel("Another Window % d" % randint(0,100))
#         layout.addWidget(self.label)

#         self.widget.setStyleSheet('QWidget { color: rgb(255, 100, 255); border: someborder; border: 20px solid green;}')

#         # self.setWindowFlag(QtCore.Qt.WindowCloseButtonHint, False)
#         # self.setWindowFlag(QtCore.Qt.WindowTitleHint, False)
#         self.setLayout(layout)
#         self.setGeometry(700, 500, 650, 100)

#         ## To remove title bar
#         self.setWindowFlags(QtCore.Qt.Window | QtCore.Qt.CustomizeWindowHint | QtCore.Qt.Tool)


#     def start_progress(self):  # To restart the progress every time
#         self.show()




### USER INTERFACE CODE ###
### =================== ###
class Ui_TabWidget(QWidget):
    def setupUi(self, TabWidget):

        # self.popup = PopUpProgressBar()

        TabWidget.setObjectName("TabWidget")
        TabWidget.resize(1405, 793)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(TabWidget.sizePolicy().hasHeightForWidth())
        TabWidget.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        TabWidget.setFont(font)
        TabWidget.setStyleSheet("QTabWidget{background-color: rgb(255, 237, 242);}\n"
"")
        self.tab = QtWidgets.QWidget()
        self.tab.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.tab.setObjectName("tab")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.tab)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")


        self.lineEdit_3 = QtWidgets.QLineEdit(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_3.sizePolicy().hasHeightForWidth())
        self.lineEdit_3.setSizePolicy(sizePolicy)
        self.lineEdit_3.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_3.setFont(font)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.gridLayout.addWidget(self.lineEdit_3, 4, 4, 1, 2)
        self.label_7 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_7.sizePolicy().hasHeightForWidth())
        self.label_7.setSizePolicy(sizePolicy)
        self.label_7.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_7.setAlignment(QtCore.Qt.AlignCenter)
        self.label_7.setObjectName("label_7")
        self.gridLayout.addWidget(self.label_7, 9, 0, 1, 1)
        self.label_8 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_8.sizePolicy().hasHeightForWidth())
        self.label_8.setSizePolicy(sizePolicy)
        self.label_8.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        self.gridLayout.addWidget(self.label_8, 9, 3, 1, 3)
        self.pushButton_11 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_11.sizePolicy().hasHeightForWidth())
        self.pushButton_11.setSizePolicy(sizePolicy)
        self.pushButton_11.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_11.setFont(font)
        self.pushButton_11.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_11.setObjectName("pushButton_11")

        #################################### Rural/Urban ##################################
        self.pushButton_11.installEventFilter(TabWidget)
        self.pushButton_11.clicked.connect(self.onSelectRuralUrban)

        self.gridLayout.addWidget(self.pushButton_11, 9, 6, 1, 1)
        self.pushButton_10 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_10.sizePolicy().hasHeightForWidth())
        self.pushButton_10.setSizePolicy(sizePolicy)
        self.pushButton_10.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_10.setFont(font)
        self.pushButton_10.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_10.setObjectName("pushButton_10")

        ####################################### Health Block ################################
        self.pushButton_10.installEventFilter(TabWidget)
        self.pushButton_10.clicked.connect(self.onSelectHealthBlock)

        self.gridLayout.addWidget(self.pushButton_10, 8, 6, 1, 1)
        self.label = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label.sizePolicy().hasHeightForWidth())
        self.label.setSizePolicy(sizePolicy)
        self.label.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet("QLabel{background-color: #003679; color : white;}\n"
"")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 10)
        self.label_10 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_10.sizePolicy().hasHeightForWidth())
        self.label_10.setSizePolicy(sizePolicy)
        self.label_10.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_10.setAlignment(QtCore.Qt.AlignCenter)
        self.label_10.setObjectName("label_10")
        self.gridLayout.addWidget(self.label_10, 10, 3, 1, 3)
        self.pushButton_14 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_14.sizePolicy().hasHeightForWidth())
        self.pushButton_14.setSizePolicy(sizePolicy)
        self.pushButton_14.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_14.setFont(font)
        self.pushButton_14.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_14.setObjectName("pushButton_14")

        #################################### Export Validated Result #############################
        self.pushButton_14.clicked.connect(self.saveValidatedData)

        self.gridLayout.addWidget(self.pushButton_14, 9, 9, 1, 1)
        self.label_19 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_19.sizePolicy().hasHeightForWidth())
        self.label_19.setSizePolicy(sizePolicy)
        self.label_19.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setItalic(True)
        self.label_19.setFont(font)
        self.label_19.setStyleSheet("QLabel{color : red;}")
        self.label_19.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_19.setObjectName("label_19")
        self.gridLayout.addWidget(self.label_19, 5, 6, 1, 4)
        self.label_5 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_5.sizePolicy().hasHeightForWidth())
        self.label_5.setSizePolicy(sizePolicy)
        self.label_5.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setStyleSheet("color: grey;background-color: rgb(255, 255, 255);")
        self.label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.label_5.setObjectName("label_5")
        self.gridLayout.addWidget(self.label_5, 8, 0, 1, 1)
        self.label_9 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_9.sizePolicy().hasHeightForWidth())
        self.label_9.setSizePolicy(sizePolicy)
        self.label_9.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_9.setAlignment(QtCore.Qt.AlignCenter)
        self.label_9.setObjectName("label_9")
        self.gridLayout.addWidget(self.label_9, 10, 0, 1, 1)
        self.pushButton_8 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_8.sizePolicy().hasHeightForWidth())
        self.pushButton_8.setSizePolicy(sizePolicy)
        self.pushButton_8.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_8.setFont(font)
        self.pushButton_8.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_8.setObjectName("pushButton_8")

        ######################################## Sub-District #####################################
        self.pushButton_8.installEventFilter(TabWidget)
        self.pushButton_8.clicked.connect(self.onSelectSubDistrict)

        self.gridLayout.addWidget(self.pushButton_8, 10, 1, 1, 2)
        self.pushButton_6 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_6.sizePolicy().hasHeightForWidth())
        self.pushButton_6.setSizePolicy(sizePolicy)
        self.pushButton_6.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_6.setFont(font)
        self.pushButton_6.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_6.setObjectName("pushButton_6")

        ################################# State ################################
        self.pushButton_6.installEventFilter(TabWidget)
        self.pushButton_6.clicked.connect(self.onSelectState)

        self.gridLayout.addWidget(self.pushButton_6, 8, 1, 1, 2)
        self.pushButton_15 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_15.sizePolicy().hasHeightForWidth())
        self.pushButton_15.setSizePolicy(sizePolicy)
        self.pushButton_15.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_15.setFont(font)
        self.pushButton_15.setStyleSheet("background-color: #B00020;color: white;")
        self.pushButton_15.setObjectName("pushButton_15")

        ######################################### Reset ####################################
        self.pushButton_15.clicked.connect(self.reset)

        self.gridLayout.addWidget(self.pushButton_15, 10, 9, 1, 1)
        self.pushButton_7 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_7.sizePolicy().hasHeightForWidth())
        self.pushButton_7.setSizePolicy(sizePolicy)
        self.pushButton_7.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_7.setFont(font)
        self.pushButton_7.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_7.setObjectName("pushButton_7")

        ########################################### District #######################################
        self.pushButton_7.installEventFilter(TabWidget)
        self.pushButton_7.clicked.connect(self.onSelectDistrict)

        self.gridLayout.addWidget(self.pushButton_7, 9, 1, 1, 2)
        self.pushButton = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton.sizePolicy().hasHeightForWidth())
        self.pushButton.setSizePolicy(sizePolicy)
        self.pushButton.setMinimumSize(QtCore.QSize(0, 0))
        self.pushButton.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("QPushButton{background-color: #73C067;color: white;}")
        self.pushButton.setObjectName("pushButton")

        ####################### Upload 
        
        self.pushButton.clicked.connect(self.get_file)


        self.gridLayout.addWidget(self.pushButton, 2, 0, 1, 1)
        self.label_6 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_6.sizePolicy().hasHeightForWidth())
        self.label_6.setSizePolicy(sizePolicy)
        self.label_6.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        self.gridLayout.addWidget(self.label_6, 8, 3, 1, 3)
        self.pushButton_5 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_5.sizePolicy().hasHeightForWidth())
        self.pushButton_5.setSizePolicy(sizePolicy)
        self.pushButton_5.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_5.setFont(font)
        # self.pushButton.setStyleSheet("QPushButton{background-color: #73C067;color: white;}")
        self.pushButton_5.setStyleSheet("QPushButton"
							"{"
							"background-color : #003679;color: white;"
							"}"
							"QPushButton::pressed"
							"{"
							"background-color : #1E88E5;"
							"}"
							)
        self.pushButton_5.setObjectName("pushButton_5")

        ###################################### Validate ################################
        self.pushButton_5.clicked.connect(self.VerifyFType)

        self.gridLayout.addWidget(self.pushButton_5, 4, 6, 1, 1)
        self.pushButton_12 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_12.sizePolicy().hasHeightForWidth())
        self.pushButton_12.setSizePolicy(sizePolicy)
        self.pushButton_12.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_12.setFont(font)
        self.pushButton_12.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_12.setObjectName("pushButton_12")

        ##################################### Ownership #####################################
        self.pushButton_12.installEventFilter(TabWidget)
        self.pushButton_12.clicked.connect(self.onSelectOwnership)

        self.gridLayout.addWidget(self.pushButton_12, 10, 6, 1, 1)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_2.sizePolicy().hasHeightForWidth())
        self.lineEdit_2.setSizePolicy(sizePolicy)
        self.lineEdit_2.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.gridLayout.addWidget(self.lineEdit_2, 4, 0, 1, 4)
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem, 1, 0, 1, 1)
        self.label_3 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_3.sizePolicy().hasHeightForWidth())
        self.label_3.setSizePolicy(sizePolicy)
        self.label_3.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(True)
        font.setWeight(50)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("QLabel{color : red;}")
        self.label_3.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_3.setObjectName("label_3")
        self.gridLayout.addWidget(self.label_3, 3, 0, 1, 7)
        self.pushButton_9 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_9.sizePolicy().hasHeightForWidth())
        self.pushButton_9.setSizePolicy(sizePolicy)
        self.pushButton_9.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_9.setFont(font)
        self.pushButton_9.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_9.setObjectName("pushButton_9")

        ################################### Block #######################################
        self.pushButton_9.installEventFilter(TabWidget)
        self.pushButton_9.clicked.connect(self.onSelectBlock)

        self.gridLayout.addWidget(self.pushButton_9, 11, 1, 1, 2)
        self.label_12 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_12.sizePolicy().hasHeightForWidth())
        self.label_12.setSizePolicy(sizePolicy)
        self.label_12.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_12.setFont(font)
        self.label_12.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.gridLayout.addWidget(self.label_12, 11, 3, 1, 3)
        self.label_11 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_11.sizePolicy().hasHeightForWidth())
        self.label_11.setSizePolicy(sizePolicy)
        self.label_11.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_11.setFont(font)
        self.label_11.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_11.setAlignment(QtCore.Qt.AlignCenter)
        self.label_11.setObjectName("label_11")
        self.gridLayout.addWidget(self.label_11, 11, 0, 1, 1)
        self.pushButton_13 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_13.sizePolicy().hasHeightForWidth())
        self.pushButton_13.setSizePolicy(sizePolicy)
        self.pushButton_13.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        self.pushButton_13.setFont(font)
        self.pushButton_13.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_13.setObjectName("pushButton_13")

        ############################################ Facility Name ###################################
        self.pushButton_13.installEventFilter(TabWidget)
        self.pushButton_13.clicked.connect(self.onSelectFacilityName)

        self.gridLayout.addWidget(self.pushButton_13, 11, 6, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem1, 9, 7, 1, 1)
        self.label_4 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_4.sizePolicy().hasHeightForWidth())
        self.label_4.setSizePolicy(sizePolicy)
        self.label_4.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label_4.setStyleSheet("QLabel{background-color: white;color: grey;}")
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setObjectName("label_4")
        self.gridLayout.addWidget(self.label_4, 6, 1, 2, 5)
        self.lineEdit = QtWidgets.QLineEdit(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit.sizePolicy().hasHeightForWidth())
        self.lineEdit.setSizePolicy(sizePolicy)
        self.lineEdit.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout.addWidget(self.lineEdit, 2, 1, 1, 6)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem2, 9, 8, 1, 1)
        self.pushButton_4 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_4.sizePolicy().hasHeightForWidth())
        self.pushButton_4.setSizePolicy(sizePolicy)
        self.pushButton_4.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setStyleSheet("QPushButton{background-color: #F47B1F;color: white;}")
        self.pushButton_4.setObjectName("pushButton_4")

        ####################################### User Manual(Button_4) #####################################
        self.pushButton_4.clicked.connect(self.UserManualEnglish)

        self.gridLayout.addWidget(self.pushButton_4, 1, 9, 1, 1)
        self.pushButton_2 = QtWidgets.QPushButton(self.tab)
        self.pushButton_2.clicked.connect(self.export)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_2.sizePolicy().hasHeightForWidth())
        self.pushButton_2.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_2.setObjectName("pushButton_2")
        self.gridLayout.addWidget(self.pushButton_2, 8, 9, 1, 1)
        self.gridLayout.setColumnStretch(0, 2)
        self.gridLayout.setColumnStretch(1, 1)
        self.gridLayout.setColumnStretch(2, 1)
        self.gridLayout.setColumnStretch(3, 1)
        self.gridLayout.setColumnStretch(4, 1)
        self.gridLayout.setColumnStretch(5, 1)
        self.gridLayout.setColumnStretch(6, 2)
        self.gridLayout.setColumnStretch(9, 2)
        self.gridLayout.setRowStretch(0, 2)
        self.gridLayout.setRowStretch(3, 1)
        self.gridLayout.setRowStretch(5, 1)
        self.gridLayout.setRowStretch(7, 2)
        self.gridLayout_2.addLayout(self.gridLayout, 0, 0, 1, 1)
        TabWidget.addTab(self.tab, "")
        self.tab1 = QtWidgets.QWidget()
        self.tab1.setObjectName("tab1")
        self.gridLayout_5 = QtWidgets.QGridLayout(self.tab1)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.gridLayout_4 = QtWidgets.QGridLayout()
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.pushButton_43 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_43.sizePolicy().hasHeightForWidth())
        self.pushButton_43.setSizePolicy(sizePolicy)
        self.pushButton_43.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_43.setFont(font)
        self.pushButton_43.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_43.setObjectName("pushButton_43")

        ######################################## Block ##################################
        self.pushButton_43.installEventFilter(TabWidget)
        self.pushButton_43.clicked.connect(self.onSelectBlock)

        self.gridLayout_4.addWidget(self.pushButton_43, 11, 1, 1, 2)
        spacerItem3 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_4.addItem(spacerItem3, 1, 2, 1, 1)
        self.label_50 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_50.sizePolicy().hasHeightForWidth())
        self.label_50.setSizePolicy(sizePolicy)
        self.label_50.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setItalic(True)
        self.label_50.setFont(font)
        self.label_50.setStyleSheet("QLabel{color : red;}")
        self.label_50.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_50.setObjectName("label_50")
        self.gridLayout_4.addWidget(self.label_50, 5, 6, 1, 4)
        self.pushButton_40 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_40.sizePolicy().hasHeightForWidth())
        self.pushButton_40.setSizePolicy(sizePolicy)
        self.pushButton_40.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_40.setFont(font)
        self.pushButton_40.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_40.setObjectName("pushButton_40")

        ##################################### District in hindi version ###########################
        self.pushButton_40.installEventFilter(TabWidget)
        self.pushButton_40.clicked.connect(self.onSelectDistrict)

        self.gridLayout_4.addWidget(self.pushButton_40, 9, 1, 1, 2)
        self.lineEdit_8 = QtWidgets.QLineEdit(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_8.sizePolicy().hasHeightForWidth())
        self.lineEdit_8.setSizePolicy(sizePolicy)
        self.lineEdit_8.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_8.setFont(font)
        self.lineEdit_8.setObjectName("lineEdit_8")
        self.gridLayout_4.addWidget(self.lineEdit_8, 4, 0, 1, 4)
        self.label_47 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_47.sizePolicy().hasHeightForWidth())
        self.label_47.setSizePolicy(sizePolicy)
        self.label_47.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_47.setFont(font)
        self.label_47.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_47.setAlignment(QtCore.Qt.AlignCenter)
        self.label_47.setObjectName("label_47")
        self.gridLayout_4.addWidget(self.label_47, 9, 3, 1, 3)
        self.pushButton_39 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_39.sizePolicy().hasHeightForWidth())
        self.pushButton_39.setSizePolicy(sizePolicy)
        self.pushButton_39.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_39.setFont(font)
        self.pushButton_39.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_39.setObjectName("pushButton_39")

        ######################################### Sub-District in Hindi Version ###############################
        self.pushButton_39.installEventFilter(TabWidget)
        self.pushButton_39.clicked.connect(self.onSelectSubDistrict)

        self.gridLayout_4.addWidget(self.pushButton_39, 10, 1, 1, 2)
        self.pushButton_41 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_41.sizePolicy().hasHeightForWidth())
        self.pushButton_41.setSizePolicy(sizePolicy)
        self.pushButton_41.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_41.setFont(font)
        self.pushButton_41.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_41.setObjectName("pushButton_41")

        ######################## Hindi Version Export Validated Data
        self.pushButton_41.clicked.connect(self.saveValidatedData)

        self.gridLayout_4.addWidget(self.pushButton_41, 9, 9, 1, 1)
        self.pushButton_35 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_35.sizePolicy().hasHeightForWidth())
        self.pushButton_35.setSizePolicy(sizePolicy)
        self.pushButton_35.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_35.setFont(font)
        self.pushButton_35.setStyleSheet("QPushButton"
							"{"
							"background-color : #003679;color: white;"
							"}"
							"QPushButton::pressed"
							"{"
							"background-color : #1E88E5;"
							"}"
							)
        self.pushButton_35.setObjectName("pushButton_35")

        ##################################### Validate in hindi version ###########################
        self.pushButton_35.clicked.connect(self.VerifyFType)

        self.gridLayout_4.addWidget(self.pushButton_35, 4, 6, 1, 1)
        self.pushButton_33 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_33.sizePolicy().hasHeightForWidth())
        self.pushButton_33.setSizePolicy(sizePolicy)
        self.pushButton_33.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_33.setFont(font)
        self.pushButton_33.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_33.setObjectName("pushButton_33")

        ########################## Hindi Version Ownership
        self.pushButton_33.installEventFilter(TabWidget)
        self.pushButton_33.clicked.connect(self.onSelectOwnership)

        self.gridLayout_4.addWidget(self.pushButton_33, 10, 6, 1, 1)
        self.label_39 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_39.sizePolicy().hasHeightForWidth())
        self.label_39.setSizePolicy(sizePolicy)
        self.label_39.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_39.setFont(font)
        self.label_39.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_39.setAlignment(QtCore.Qt.AlignCenter)
        self.label_39.setObjectName("label_39")
        self.gridLayout_4.addWidget(self.label_39, 9, 0, 1, 1)
        self.label_52 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_52.sizePolicy().hasHeightForWidth())
        self.label_52.setSizePolicy(sizePolicy)
        self.label_52.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_52.setFont(font)
        self.label_52.setStyleSheet("color: grey;background-color: rgb(255, 255, 255);")
        self.label_52.setAlignment(QtCore.Qt.AlignCenter)
        self.label_52.setObjectName("label_52")
        self.gridLayout_4.addWidget(self.label_52, 8, 0, 1, 1)
        self.pushButton_42 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_42.sizePolicy().hasHeightForWidth())
        self.pushButton_42.setSizePolicy(sizePolicy)
        self.pushButton_42.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_42.setFont(font)
        self.pushButton_42.setStyleSheet("background-color: #B00020;color: white;")
        self.pushButton_42.setObjectName("pushButton_42")

        ######################## Hindi Version reset
        self.pushButton_42.clicked.connect(self.reset)

        self.gridLayout_4.addWidget(self.pushButton_42, 10, 9, 1, 1)
        self.lineEdit_7 = QtWidgets.QLineEdit(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_7.sizePolicy().hasHeightForWidth())
        self.lineEdit_7.setSizePolicy(sizePolicy)
        self.lineEdit_7.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setFamily("Agency FB")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_7.setFont(font)
        self.lineEdit_7.setObjectName("lineEdit_7")
        self.gridLayout_4.addWidget(self.lineEdit_7, 4, 4, 1, 2)
        self.pushButton_32 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_32.sizePolicy().hasHeightForWidth())
        self.pushButton_32.setSizePolicy(sizePolicy)
        self.pushButton_32.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_32.setFont(font)
        self.pushButton_32.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_32.setObjectName("pushButton_32")

        ################################ Hindi Version Rural/Urban
        self.pushButton_32.installEventFilter(TabWidget)
        self.pushButton_32.clicked.connect(self.onSelectRuralUrban)

        self.gridLayout_4.addWidget(self.pushButton_32, 9, 6, 1, 1)
        self.pushButton_44 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_44.sizePolicy().hasHeightForWidth())
        self.pushButton_44.setSizePolicy(sizePolicy)
        self.pushButton_44.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_44.setFont(font)
        self.pushButton_44.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_44.setObjectName("pushButton_44")

        ##################################### State in hindi version ###########################
        self.pushButton_44.installEventFilter(TabWidget)
        self.pushButton_44.clicked.connect(self.onSelectState)

        self.gridLayout_4.addWidget(self.pushButton_44, 8, 1, 1, 2)
        self.label_41 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_41.sizePolicy().hasHeightForWidth())
        self.label_41.setSizePolicy(sizePolicy)
        self.label_41.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_41.setFont(font)
        self.label_41.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_41.setAlignment(QtCore.Qt.AlignCenter)
        self.label_41.setObjectName("label_41")
        self.gridLayout_4.addWidget(self.label_41, 10, 0, 1, 1)
        self.pushButton_37 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_37.sizePolicy().hasHeightForWidth())
        self.pushButton_37.setSizePolicy(sizePolicy)
        self.pushButton_37.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_37.setFont(font)
        self.pushButton_37.setStyleSheet("QPushButton{background-color: #73C067;color: white;}")
        self.pushButton_37.setObjectName("pushButton_37")

        ################################### Upload in hindi version ###################################
        self.pushButton_37.clicked.connect(self.get_file)

        self.gridLayout_4.addWidget(self.pushButton_37, 2, 0, 1, 1)
        self.label_56 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_56.sizePolicy().hasHeightForWidth())
        self.label_56.setSizePolicy(sizePolicy)
        self.label_56.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setItalic(True)
        self.label_56.setFont(font)
        self.label_56.setStyleSheet("QLabel{color : red;}")
        self.label_56.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_56.setObjectName("label_56")
        self.gridLayout_4.addWidget(self.label_56, 3, 0, 1, 7)
        self.label_53 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_53.sizePolicy().hasHeightForWidth())
        self.label_53.setSizePolicy(sizePolicy)
        self.label_53.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_53.setFont(font)
        self.label_53.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_53.setAlignment(QtCore.Qt.AlignCenter)
        self.label_53.setObjectName("label_53")
        self.gridLayout_4.addWidget(self.label_53, 10, 3, 1, 3)
        self.label_46 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_46.sizePolicy().hasHeightForWidth())
        self.label_46.setSizePolicy(sizePolicy)
        self.label_46.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_46.setFont(font)
        self.label_46.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_46.setAlignment(QtCore.Qt.AlignCenter)
        self.label_46.setObjectName("label_46")
        self.gridLayout_4.addWidget(self.label_46, 8, 3, 1, 3)
        self.pushButton_38 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_38.sizePolicy().hasHeightForWidth())
        self.pushButton_38.setSizePolicy(sizePolicy)
        self.pushButton_38.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_38.setFont(font)
        self.pushButton_38.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_38.setObjectName("pushButton_38")

        ################################## Hindi Version Health Block
        self.pushButton_38.installEventFilter(TabWidget)
        self.pushButton_38.clicked.connect(self.onSelectHealthBlock)

        self.gridLayout_4.addWidget(self.pushButton_38, 8, 6, 1, 1)
        self.label_55 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_55.sizePolicy().hasHeightForWidth())
        self.label_55.setSizePolicy(sizePolicy)
        self.label_55.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label_55.setFont(font)
        self.label_55.setStyleSheet("QLabel{background-color: #003679; color : white;}\n"
"")
        self.label_55.setAlignment(QtCore.Qt.AlignCenter)
        self.label_55.setObjectName("label_55")
        self.gridLayout_4.addWidget(self.label_55, 0, 0, 1, 10)
        self.label_40 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_40.sizePolicy().hasHeightForWidth())
        self.label_40.setSizePolicy(sizePolicy)
        self.label_40.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Agency FB")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_40.setFont(font)
        self.label_40.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_40.setAlignment(QtCore.Qt.AlignCenter)
        self.label_40.setObjectName("label_40")
        self.gridLayout_4.addWidget(self.label_40, 11, 0, 1, 1)
        self.label_54 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.label_54.sizePolicy().hasHeightForWidth())
        self.label_54.setSizePolicy(sizePolicy)
        self.label_54.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_54.setFont(font)
        self.label_54.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_54.setAlignment(QtCore.Qt.AlignCenter)
        self.label_54.setObjectName("label_54")
        self.gridLayout_4.addWidget(self.label_54, 11, 3, 1, 3)
        self.pushButton_36 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_36.sizePolicy().hasHeightForWidth())
        self.pushButton_36.setSizePolicy(sizePolicy)
        self.pushButton_36.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.pushButton_36.setFont(font)
        self.pushButton_36.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_36.setObjectName("pushButton_36")

        ############################## Hindi Version Ownsership
        self.pushButton_36.installEventFilter(TabWidget)
        self.pushButton_36.clicked.connect(self.onSelectFacilityName)

        self.gridLayout_4.addWidget(self.pushButton_36, 11, 6, 1, 1)
        spacerItem4 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_4.addItem(spacerItem4, 9, 7, 1, 1)
        self.label_48 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_48.sizePolicy().hasHeightForWidth())
        self.label_48.setSizePolicy(sizePolicy)
        self.label_48.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_48.setFont(font)
        self.label_48.setStyleSheet("QLabel{background-color: white;color: grey;}")
        self.label_48.setAlignment(QtCore.Qt.AlignCenter)
        self.label_48.setObjectName("label_48")
        self.gridLayout_4.addWidget(self.label_48, 6, 1, 2, 5)
        spacerItem5 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_4.addItem(spacerItem5, 9, 8, 1, 1)
        self.lineEdit_9 = QtWidgets.QLineEdit(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_9.sizePolicy().hasHeightForWidth())
        self.lineEdit_9.setSizePolicy(sizePolicy)
        self.lineEdit_9.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_9.setFont(font)
        self.lineEdit_9.setObjectName("lineEdit_9")
        self.gridLayout_4.addWidget(self.lineEdit_9, 2, 1, 1, 6)
        self.pushButton_45 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.pushButton_45.sizePolicy().hasHeightForWidth())
        self.pushButton_45.setSizePolicy(sizePolicy)
        self.pushButton_45.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton_45.setFont(font)
        self.pushButton_45.setStyleSheet("QPushButton{background-color: #F47B1F;color: white;}")
        self.pushButton_45.setObjectName("pushButton_45")

        ##################################### User Manual in hindi version ###########################
        self.pushButton_45.clicked.connect(self.UserManualHindi)

        self.gridLayout_4.addWidget(self.pushButton_45, 1, 9, 1, 1)
        self.pushButton_3 = QtWidgets.QPushButton(self.tab1)
        self.pushButton_3.clicked.connect(self.export)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_3.sizePolicy().hasHeightForWidth())
        self.pushButton_3.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_3.setObjectName("pushButton_3")
        self.gridLayout_4.addWidget(self.pushButton_3, 8, 9, 1, 1)
        self.gridLayout_4.setColumnStretch(0, 2)
        self.gridLayout_4.setColumnStretch(1, 1)
        self.gridLayout_4.setColumnStretch(2, 1)
        self.gridLayout_4.setColumnStretch(3, 1)
        self.gridLayout_4.setColumnStretch(4, 1)
        self.gridLayout_4.setColumnStretch(5, 1)
        self.gridLayout_4.setColumnStretch(6, 2)
        self.gridLayout_4.setColumnStretch(9, 2)
        self.gridLayout_4.setRowStretch(0, 2)
        self.gridLayout_4.setRowStretch(1, 1)
        self.gridLayout_4.setRowStretch(2, 1)
        self.gridLayout_4.setRowStretch(3, 1)
        self.gridLayout_4.setRowStretch(5, 1)
        self.gridLayout_4.setRowStretch(7, 2)
        self.gridLayout_5.addLayout(self.gridLayout_4, 0, 0, 1, 1)
        TabWidget.addTab(self.tab1, "")

        self.retranslateUi(TabWidget)
        TabWidget.setCurrentIndex(0)

        # ProgressBar addition
        # --------------------
        # self.popup = PopUpProgressBar()

        QtCore.QMetaObject.connectSlotsByName(TabWidget)


    ##! ********************************************************************************************
    ##! Use this function To attach files to the exe file (eg - png, txt, jpg etc) using pyinstaller
    ##! ********************************************************************************************
    def resource_path(self, relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)


    def retranslateUi(self, TabWidget):
        _translate = QtCore.QCoreApplication.translate
        TabWidget.setWindowTitle(_translate("TabWidget", "ADVTool"))
        self.lineEdit_3.setPlaceholderText(_translate("TabWidget", " Month, Year"))
        self.label_7.setText(_translate("TabWidget", "District"))
        self.label_8.setText(_translate("TabWidget", "Rural / Urban"))
        self.pushButton_11.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_10.setText(_translate("TabWidget", "-- All Selected --"))
        self.label.setText(_translate("TabWidget", "Data Validation Tool"))
        self.label_10.setText(_translate("TabWidget", "Ownership"))
        self.pushButton_14.setText(_translate("TabWidget", "Download Validated Data"))
        self.label_19.setText(_translate("TabWidget", "* Press Validate button to perform validation check on your data"))
        self.label_5.setText(_translate("TabWidget", "State"))
        self.label_9.setText(_translate("TabWidget", "Sub-District"))
        self.pushButton_8.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_6.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_15.setText(_translate("TabWidget", "Reset"))
        self.pushButton_7.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton.setText(_translate("TabWidget", "Upload"))
        self.label_6.setText(_translate("TabWidget", "Health Block"))
        self.pushButton_5.setText(_translate("TabWidget", "Validate"))
        self.pushButton_12.setText(_translate("TabWidget", "-- All Selected --"))
        self.lineEdit_2.setPlaceholderText(_translate("TabWidget", "  Facility Type selected will display here ..."))
        self.label_3.setText(_translate("TabWidget", "* Upload data in .xls / .xlsx format for one month and one facility type only."))
        self.pushButton_9.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_12.setText(_translate("TabWidget", "Facility Name"))
        self.label_11.setText(_translate("TabWidget", "Block"))
        self.pushButton_13.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_4.setText(_translate("TabWidget", "Select Filters"))
        self.lineEdit.setPlaceholderText(_translate("TabWidget", "  Your uploaded file name will display here ..."))
        self.pushButton_4.setText(_translate("TabWidget", "User Manual"))
        self.pushButton_2.setText(_translate("TabWidget", " Download Validated Results "))
        TabWidget.setTabText(TabWidget.indexOf(self.tab), _translate("TabWidget", "English Version"))
        self.pushButton_43.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_50.setText(_translate("TabWidget", "* अपने डेटा पर सत्यापन जांच करने के लिए मान्य बटन दबाएं"))
        self.pushButton_40.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.lineEdit_8.setPlaceholderText(_translate("TabWidget", "चयनित स्वास्थ्य केंद्र प्रकार यहां प्रदर्शित होगा..."))
        self.label_47.setText(_translate("TabWidget", "ग्रामीण/शहरी"))
        self.pushButton_39.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.pushButton_41.setText(_translate("TabWidget", "डाउनलोड मान्य डेटा"))
        self.pushButton_35.setText(_translate("TabWidget", "पुष्टि"))
        self.pushButton_33.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_39.setText(_translate("TabWidget", "जिला"))
        self.label_52.setText(_translate("TabWidget", "राज्य"))
        self.pushButton_42.setText(_translate("TabWidget", "रीसेट"))
        self.lineEdit_7.setPlaceholderText(_translate("TabWidget", "महीना,  वर्ष"))
        self.pushButton_32.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.pushButton_44.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_41.setText(_translate("TabWidget", "उप - जिला"))
        self.pushButton_37.setText(_translate("TabWidget", "अपलोड"))
        self.label_56.setText(_translate("TabWidget", "* केवल एक प्रकार के स्वास्थ्य केंद्र और केवल एक महीने  के लिए .xls / .xlsx प्रारूप में डेटा अपलोड करें।"))
        self.label_53.setText(_translate("TabWidget", "स्वामित्व"))
        self.label_46.setText(_translate("TabWidget", "स्वास्थ्य प्रखंड"))
        self.pushButton_38.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_55.setText(_translate("TabWidget", "डाटा वेलिडेशन टूल"))
        self.label_40.setText(_translate("TabWidget", "प्रखंड"))
        self.label_54.setText(_translate("TabWidget", "स्वास्थ्य केंद्र का नाम"))
        self.pushButton_36.setText(_translate("TabWidget", "-- सभी चयनित --"))
        self.label_48.setText(_translate("TabWidget", "फ़िल्टर का चयन करें"))
        self.lineEdit_9.setPlaceholderText(_translate("TabWidget", "आपकी अपलोड की गई फ़ाइल का नाम यहां प्रदर्शित होगा..."))
        self.pushButton_45.setText(_translate("TabWidget", "यूजर मैनुअल"))
        self.pushButton_3.setText(_translate("TabWidget", "डाउनलोड रिजलट"))
        TabWidget.setTabText(TabWidget.indexOf(self.tab1), _translate("TabWidget", "Hindi Version"))


    ##! ******************************************************************************************** 
    ##! Use this function To attach files to the exe file (eg - png, txt, jpg etc) using pyinstaller
    ##! ********************************************************************************************
    def resource_path(self, relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)


    ''' Upload Function '''
    # =====================

    def get_file(self):
        global fileName, df_, len_df_SummReport, new_header, items_date, unique_list

        ### Calling upload UI 
        click_Upload = Ui_Dialog_Upload()
        click_Upload.exec_()

        # try:

        ### Check for empty files
        if click_Upload.fileName[0] != "":
            
            if len(click_Upload.fileName) == 1:
                # Read uploaded excel file
                df_ = pd.read_excel(click_Upload.fileName[0])

                # Dropping last two rows
                df_.drop(df_.index[[-1, -2]], inplace=True)

                # Extracting string from 1st cell of dataframe
                str_to_extr_MonthYear = str(df_.iloc[0])

                # grab the first row for the header
                new_header = df_.iloc[0]

                # #take the data less the header row
                df_ = df_[0: -1]

                # set the header row as the df header
                df_.columns = new_header

                # Extracting Month , Year from string
                results = re.findall(
                    r"[abceglnoprtuvyADFJMNOS|]{3}[\s-]\d{2,4}", str_to_extr_MonthYear)

                l = df_.columns.values
                cnt = 0
                for i in l:
                    if i != 'M1 [Ante Natal Care (ANC)]':
                        cnt += 1
                    else:
                        break

                # Partial list of headers
                lst1 = df_.columns[:cnt].values

                # Picking row items after 18th row to merge with lst1
                lst2 = df_.iloc[1, cnt:].values.tolist()

                # Merging both lists
                lst3 = np.concatenate((lst1, lst2))

                # Assign lst3 as new column header
                df_.columns = lst3

                # Taking DataFrame from second row
                df_ = df_[3:]

                # Reindexing dataframe
                df_ = df_.reset_index(drop=True)

                df_ = df_.iloc[:, 1:]

                ###! --- Removing Duplicate Columns ---
                df_ = df_.loc[~df_.index.duplicated(keep='first')]

                ### Add a new column as date in the dataframe
                df_['DATE'] = results[0]

                # df_ = df_.set_index('DATE')

                len_df_SummReport = len(df_.columns)

                # Disabling upload Button
                self.pushButton.setDisabled(True)

                # list_set = df_['Facility Type'].tolist()
                # unique_list = list(set(list_set))

                list_set = df_['Facility Type'].value_counts()
                unique_list = list_set.idxmax()


            elif len(click_Upload.fileName) > 1:
                # list of excel files we want to merge.
                # pd.read_excel(file_path) reads the
                # excel data into pandas dataframe.

                for i in range(len(click_Upload.fileName)):

                    if i == 0:

                        df_ = pd.read_excel(click_Upload.fileName[0])    ### File reading

                        # Dropping last two rows
                        df_.drop(df_.index[[-1, -2]], inplace=True)

                        # Extracting string from 1st cell of dataframe
                        str_to_extr_MonthYear = str(df_.iloc[0])

                        # grab the first row for the header
                        new_header = df_.iloc[0]

                        # #take the data less the header row
                        df_ = df_[0: -1]

                        # set the header row as the df header
                        df_.columns = new_header

                        # Extracting Month , Year from string
                        results = re.findall(
                            r"[abceglnoprtuvyADFJMNOS|]{3}[\s-]\d{2,4}", str_to_extr_MonthYear)


                        l = df_.columns.values
                        cnt = 0
                        for i in l:
                            if i != 'M1 [Ante Natal Care (ANC)]':
                                cnt += 1
                            else:
                                break

                        # Partial list of headers
                        lst1 = df_.columns[:cnt].values

                        # Picking row items after 18th row to merge with lst1
                        lst2 = df_.iloc[1, cnt:].values.tolist()

                        # Merging both lists
                        lst3 = np.concatenate((lst1, lst2))

                        # Assign lst3 as new column header
                        df_.columns = lst3

                        # Taking DataFrame from second row
                        df_ = df_[3:]

                        # Reindexing dataframe
                        df_ = df_.reset_index(drop=True)

                        df_ = df_.iloc[:, 1:]

                        ###! --- Removing Duplicate Columns ---
                        # df_ = df_.loc[~df_.index.duplicated(keep='first')]
                        df_ = df_.loc[:,~df_.columns.duplicated()].copy()

                        ### Add a new column as date in the dataframe
                        df_['DATE'] = results[0]

                        try:
                            df_.rename(columns={df_.filter(regex='^([S][u][b][-][D][i][s][t][r][i][c][t]) .*([N][a][m][e])|^ +([S][u][b][-][D][i][s][t][r][i][c][t]).* ([N][a][m][e])|^([S][u][b][-][D][i][s][t][r][i][c][t]) .*([N][a][m][e])|^ ([S][u][b][-][D][i][s][t][r][i][c][t]).* ([N][a][m][e])/i').columns[0]: 'Sub-District Name',},inplace=True)
                        except:
                            df_.rename(columns={df_.filter(regex='^([S][u][b][-][D][i][v][i][s][i][o][n]) .*([N][a][m][e])|^ +([S][u][b][-][D][i][v][i][s][i][o][n]).* ([N][a][m][e])|^([S][u][b][-][D][i][v][i][s][i][o][n]) .*([N][a][m][e])|^ ([S][u][b][-][D][i][v][i][s][i][o][n]).* ([N][a][m][e])/i').columns[0]: 'Sub-District Name',},inplace=True)
                        finally:
                            pass

                        # df_ = df_.set_index('DATE')

                        # Disabling upload Button
                        self.pushButton.setDisabled(True)

                        list_set = df_['Facility Type'].tolist()
                        unique_list = list(set(list_set))


                    elif i > 0:
                        print('---------------- i > 0 ------------------------')
                    
                        dff_ = pd.read_excel(click_Upload.fileName[i])

                        try:
                            # Dropping last two rows
                            dff_.drop(dff_.index[[-1, -2]], inplace=True)
                        except:
                            pass

                        # Extracting string from 1st cell of dataframe
                        str_to_extr_MonthYear = str(dff_.iloc[0])

                        # grab the first row for the header
                        new_header = dff_.iloc[0]

                        # #take the data less the header row
                        dff_ = dff_[0: -1]

                        # set the header row as the df header
                        dff_.columns = new_header

                        # Extracting Month , Year from string
                        results = re.findall(
                            r"[abceglnoprtuvyADFJMNOS|]{3}[\s-]\d{2,4}", str_to_extr_MonthYear)

                        l = dff_.columns.values
                        cnt = 0
                        for i in l:
                            if i != 'M1 [Ante Natal Care (ANC)]':
                                cnt += 1
                            else:
                                break

                        # Partial list of headers
                        lst1 = dff_.columns[:cnt].values

                        # Picking row items after 18th row to merge with lst1
                        lst2 = dff_.iloc[1, cnt:].values.tolist()

                        # Merging both lists
                        lst3 = np.concatenate((lst1, lst2))

                        # Assign lst3 as new column header
                        dff_.columns = lst3

                        # Taking DataFrame from second row
                        dff_ = dff_[3:]

                        # Reindexing dataframe
                        dff_ = dff_.reset_index(drop=True)

                        dff_ = dff_.iloc[:, 1:]

                        ###! --- Removing Duplicate Columns ---
                        # dff_ = dff_.loc[~dff_.index.duplicated(keep='first')]
                        dff_ = dff_.loc[:,~dff_.columns.duplicated()].copy()

                        ### Add a new column as date in the dataframe
                        dff_['DATE'] = results[0]

                        try:
                            dff_.rename(columns={dff_.filter(regex='^([S][u][b][-][D][i][s][t][r][i][c][t]) .*([N][a][m][e])|^ +([S][u][b][-][D][i][s][t][r][i][c][t]).* ([N][a][m][e])|^([S][u][b][-][D][i][s][t][r][i][c][t]) .*([N][a][m][e])|^ ([S][u][b][-][D][i][s][t][r][i][c][t]).* ([N][a][m][e])/i').columns[0]: 'Sub-District Name',},inplace=True)
                        except:
                            dff_.rename(columns={dff_.filter(regex='^([S][u][b][-][D][i][v][i][s][i][o][n]) .*([N][a][m][e])|^ +([S][u][b][-][D][i][v][i][s][i][o][n]).* ([N][a][m][e])|^([S][u][b][-][D][i][v][i][s][i][o][n]) .*([N][a][m][e])|^ ([S][u][b][-][D][i][v][i][s][i][o][n]).* ([N][a][m][e])/i').columns[0]: 'Sub-District Name',},inplace=True)
                        finally:
                            pass

                        # dff_ = dff_.set_index('DATE') 

                        # list_set2 = dff_['Facility Type'].tolist()
                        # unique_list2 = list(set(list_set2))

                        list_set2 = dff_['Facility Type'].value_counts()
                        unique_list2 = list_set2.idxmax()

                        if unique_list[0] == unique_list2[0]:
                            #### merging all the data one by one
                            df_ = df_.append(dff_)              
                            len_df_SummReport =  len(df_.columns)
                            

                        elif unique_list[0] != unique_list2[0]:
                            msgg = " can't be merged because it belongs to other facility type."
                            self.msg = QMessageBox()
                            # Set the information icon
                            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                            # Set the main message
                            self.msg.setText(msgg)
                            # Set the title of the window
                            self.msg.setWindowTitle("Success")
                            # Display the message box
                            self.msg.show()
                        else:
                            pass


            # Reindexing dataframe
            df_ = df_.reset_index(drop=True)

            # shift column 'Name' to first position
            first_column = df_.pop('DATE')
            
            # insert column using insert(position,column_name,
            # first_column) function
            df_.insert(0, 'DATE', first_column)

            list_date = df_['DATE'].to_numpy()

            items_date = str(set(list_date)).strip("{}")

            # try:
            # Filling Facility Name selected in English Version
            # self.lineEdit_2.setText(["{0}".format(col) for col in unique_list])
            self.lineEdit_2.setText(unique_list)
            # Filling Month, Year in English Version
            self.lineEdit_3.setText(items_date)
            # Filling Facility Name selected in Hindiy Version
            self.lineEdit_8.setText(unique_list)
            # Filling Month, Year in Hindi Version
            self.lineEdit_7.setText(items_date)

            # except:
            #     # Filling Facility Name selected in English Version
            #     self.lineEdit_2.setText(["{0}".format(col) for col in unique_list])
            #     # Filling Month, Year in English Version
            #     self.lineEdit_3.setText(items_date)
            #     # Filling Facility Name selected in Hindi Version
            #     self.lineEdit_8.setText(["{0}".format(col) for col in unique_list])
            #     # Filling Month, Year in Hindi Version
            #     self.lineEdit_7.setText(items_date)

            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("The file has been uploaded successfully. \n\n फ़ाइल सफलतापूर्वक अपलोड कर दी गई है।")
            # Set the title of the window
            self.msg.setWindowTitle("Success")
            # Display the message box
            self.msg.show()

        # except:
        #     pass

    

    # Upload file button functionality
    # ================================
    def loadFile(self, df_):
        return df_

    # Filtering Facility Type
    # =======================
    def VerifyFType(self):
        global df, FType
        # self.popup.start_progress()

        # try:

        FType = self.lineEdit_2.text()

        print(FType)

        if (FType == 'HWC-PHC/PHC'):
            # Signaling PHC_Validate function i.e function where validation checks are present
            df = PHC_Validate(self, df_)

        elif (FType == 'HWC-SC/SC' ):
            # Signaling HSC_Validate function i.e function where validation checks are present
            df = HSC_Validate(self, df_)

        elif (FType == 'District Hospital'):
            df = DH_Validate(self, df_)

        elif (FType == 'Sub District Hospital'):
            df = SDH_Validate(self, df_)

        elif (FType == 'Community Health Centre'):
            df = CHC_Validate(self, df_)

        else:
            raise Exception('Facility Type Name is not matching')

        # except:
        #     # Create the messagebox object
        #     self.msg = QMessageBox()
        #     # Set the information icon
        #     self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
        #     self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:#FF6347; font-family: Arial; font-size:20px;} ")
        #     # Set the main message
        #     self.msg.setText("Please upload the data files!")
        #     # Set the title of the window
        #     self.msg.setWindowTitle("Warning!")
        #     # Display the message box
        #     self.msg.show()


    '''
    # Filter to decide which filter button user clicked
    # =================================================
    '''
    def eventFilter(self, target, event):
        if target == self.pushButton_6 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_6.clicked.connect(self.onSelectState)
            return True

        elif target == self.pushButton_7 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_7.clicked.connect(self.onSelectDistrict)
            return True

        elif target == self.pushButton_13 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_13.clicked.connect(self.onSelectFacilityName)
            return True

        elif target == self.pushButton_11 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_11.clicked.connect(self.onSelectRuralUrban)
            return True

        elif target == self.pushButton_12 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_12.clicked.connect(self.onSelectOwnership)
            return True

        elif target == self.pushButton_8 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_8.clicked.connect(self.onSelectSubDistrict)
            return True

        elif target == self.pushButton_9 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_9.clicked.connect(self.onSelectBlock)
            return True

        elif target == self.pushButton_10 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_10.clicked.connect(self.onSelectHealthBlock)
            return True

        return False


    ################################################################################
    # Filter State Functionality
    
    def onSelectState(self, index):
        try:
            global final_list

            # list storing Facility Name data
            df['State'].fillna('Blank', inplace = True)

            list_set = df['State'].to_numpy()

            sorted_list = sorted(list_set, key=str.upper)

            item = set(sorted_list)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('State', item, checked=True)

            try:
                if form.exec_() == QtWidgets.QDialog.Accepted:
                    final_list = [str(s) for s in form.choices]
                elif form.exec_() == QtWidgets.QDialog.Rejected:
                    final_list = [str(s) for s in form.choices]
            except:
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            
            try :
                self.filterdataState(final_list)
            except :
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            finally:
                pass

            
        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("State column is not available in uploaded data \n\n अपलोड किए गए डेटा में स्टेट कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass
        

    # Filter data columnwise
    def filterdataState(self, final_list):
        global df
        df = df[df['State'].isin(final_list)]
        return df
    

    ################################################################################
    # Filter District Functionality

    def onSelectDistrict(self, index):
        try:
            global final_list

            # list storing Facility Name data
            df['District Name'].fillna('Blank', inplace = True)

            list_set = df['District Name'].to_numpy()

            sorted_list = sorted(list_set, key=str.upper)

            item = set(sorted_list)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('District', item, checked=True)

            try:
                if form.exec_() == QtWidgets.QDialog.Accepted:
                    final_list = [str(s) for s in form.choices]
                elif form.exec_() == QtWidgets.QDialog.Rejected:
                    final_list = [str(s) for s in form.choices]
            except:
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            
            try:
                self.filterdataDistrict(final_list)
            except :
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            finally:
                pass

            
        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("District column is not available in uploaded data \n\n अपलोड किए गए डेटा में जिला कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # Filter data columnwise
    def filterdataDistrict(self, final_list):
        global df
        df = df[df['District Name'].isin(final_list)]
        return df


    ################################################################################
    # Filter Sub District Functionality

    def onSelectSubDistrict(self, index):
        try:
            global final_list

            # list storing Facility Name data
            df['Sub-District Name'].fillna('Blank', inplace = True)

            list_set = df['Sub-District Name'].to_numpy()

            # sorted_list = sorted(list_set, key=str.upper)

            item = set(list_set)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('Sub-District Name', item, checked=True)

            try:
                if form.exec_() == QtWidgets.QDialog.Accepted:
                    final_list = [str(s) for s in form.choices]
                elif form.exec_() == QtWidgets.QDialog.Rejected:
                    final_list = [str(s) for s in form.choices]
            except:
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            
            try:
                self.filterdataSubDistrict(final_list)
            except :
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            finally:
                pass

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Sub-District column is not available in uploaded data \n\n अपलोड किए गए डेटा में उप-जिला कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # Filter data columnwise
    def filterdataSubDistrict(self, final_list):
        global df
        df = df[df['Sub-District Name'].isin(final_list)]
        return df


    ################################################################################
    # Filter Block Functionality

    def onSelectBlock(self, index):
        try:
            global final_list

            # list storing Facility Name data
            df['Block Name'].fillna('Blank', inplace = True)

            list_set = df['Block Name'].to_numpy()

            sorted_list = sorted(list_set, key=str.upper)

            item = set(sorted_list)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('Block', item, checked=True)

            try:
                if form.exec_() == QtWidgets.QDialog.Accepted:
                    final_list = [str(s) for s in form.choices]
                elif form.exec_() == QtWidgets.QDialog.Rejected:
                    final_list = [str(s) for s in form.choices]
            except:
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            
            try:
                self.filterdataBlock(final_list)
            except :
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            finally:
                pass

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Block column is not available in uploaded data \n\n अपलोड किए गए डेटा में ब्लॉक कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass


    # Filter data columnwise
    def filterdataBlock(self, final_list):
        global df
        df = df[df['Block Name'].isin(final_list)]
        return df


    ################################################################################
    # Filter Health Block Functionality

    def onSelectHealthBlock(self, index):
        try : 
            global final_list

            # list storing Facility Name data
            df['Health Block Name'].fillna('Blank', inplace = True)

            list_set = df['Health Block Name'].to_numpy()

            sorted_list = sorted(list_set, key=str.upper)

            item = set(sorted_list)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('Health Block Name', item, checked=True)

            try:
                if form.exec_() == QtWidgets.QDialog.Accepted:
                    final_list = [str(s) for s in form.choices]
                elif form.exec_() == QtWidgets.QDialog.Rejected:
                    final_list = [str(s) for s in form.choices]
            except:
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            
            try:
                self.filterdataHealthBlock(final_list)
            except :
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            finally:
                pass

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Health Block column is not available in uploaded data \n\n अपलोड किए गए डेटा में हेल्थ ब्लॉक कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass


    # Filter data columnwise
    def filterdataHealthBlock(self, final_list):
        global df
        df = df[df['Health Block Name'].isin(final_list)]
        return df


    ################################################################################
    # Filter Facility Name

    # Filter FacilityName Functionality

    def onSelectFacilityName(self, index):
        
        try:      
            # list storing Facility Name data
            df['Facility Name'].fillna('Blank', inplace = True)

            list_set = df['Facility Name'].to_numpy()

            sorted_list = sorted(list_set, key=str.upper)

            item = set(sorted_list)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('Facility Name', item, checked=True)

            try:
                if form.exec_() == QtWidgets.QDialog.Accepted:
                    final_list = [str(s) for s in form.choices]
                elif form.exec_() == QtWidgets.QDialog.Rejected:
                    final_list = [str(s) for s in form.choices]
            except:
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            
            try:
                self.filterdataFacilityName(final_list)
            except :
                pass

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Facility Name column is not available in uploaded data \n\n अपलोड किए गए डेटा में Facility Name कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # Filter data columnwise
    def filterdataFacilityName(self, final_list):
        global df
        df = df[df['Facility Name'].isin(final_list)]
        return df


    ################################################################################
    # Filter Rural/Urban

    def onSelectRuralUrban(self, index):
        try:
            global final_list
            # try:

            # list storing Facility Name data
            df['Rural/Urban'].fillna('Blank', inplace = True)

            list_set = df['Rural/Urban'].to_numpy()

            sorted_list = sorted(list_set, key=str.upper)

            item = set(sorted_list)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('Rural / Urban', item, checked=True)
            if form.exec_() == QtWidgets.QDialog.Accepted:
                final_list = [str(s) for s in form.choices]
            elif form.exec_() == QtWidgets.QDialog.Rejected:
                final_list = [str(s) for s in form.choices]
            
            try:
                self.filterdataRuralUrban(final_list)
            except :
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            finally:
                pass

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Rural/ Urban column is not available in uploaded data \n\n अपलोड किए गए डेटा में ग्रामीण/शहरी कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass
    

    # Filter data columnwise
    def filterdataRuralUrban(self, final_list):
        global df
        df = df[df['Rural/Urban'].isin(final_list)]
        return df

    ################################################################################
    # Select Ownership

    # Select Ownership Filter

    def onSelectOwnership(self, index):
        try:
            global final_list
            # try:

            # list storing Facility Name data
            df['Ownership'].fillna('Blank', inplace = True)

            list_set = df['Ownership'].to_numpy()

            sorted_list = sorted(list_set, key=str.upper)

            item = set(sorted_list)

            item = list(item)

            # app_2 = QtWidgets.QApplication(sys.argv)
            form = ChecklistDialog('Ownership', item, checked=True)
            if form.exec_() == QtWidgets.QDialog.Accepted:
                final_list = [str(s) for s in form.choices]
            elif form.exec_() == QtWidgets.QDialog.Rejected:
                final_list = [str(s) for s in form.choices]
            
            try:
                self.filterdataOwnership(final_list)
            except :
                # Create the messagebox object
                self.msg = QMessageBox()
                # Set the information icon
                self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
                self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
                # Set the main message
                self.msg.setText("Please select something. \n\n कृपया कुछ चुनें।")
                # Set the title of the window
                self.msg.setWindowTitle(" ")
                # Display the message box
                self.msg.show()
            finally:
                pass

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Ownership column is not available in uploaded data \n\n अपलोड किए गए डेटा में स्वामित्व कॉलम उपलब्ध नहीं है")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass


    # Filter data columnwise
    def filterdataOwnership(self, final_list):
        global df
        df = df[df['Ownership'].isin(final_list)]
        return df


    # To count summary of the Modified Checks
    # =======================================
    def indicator_Description(self, df):
        global df_SummReport, val_Description, UI_Val
        FType = self.lineEdit_2.text()

        # For Health Sub Centre
        if FType == 'HWC-SC/SC': 
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 200th
            val_Description = [f"Cond{i}" for i in range(53)]
            # val_Description = [
            #                     'Number of mothers provided full course of 180 IFA tablets after delivery <= Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) + Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+ Number of Institutional Deliveries conducted', 
            #                     'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC', 
            #                     'Number of PW tested using POC test for Syphilis<=Total number of pregnant women registered for ANC',
            #                     'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
            #                     'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
            #                     'Number of newborns received 6 HBNC  visits after Institutional Delivery<=Number of Institutional Deliveries conducted',
            #                     'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted, Child immunisation - Vitamin K1 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                     'Child immunisation - Vitamin K1 (Birth Dose)<=Live Birth - Male+Live Birth - Female','Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                     'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                     'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose' ]

        # For Primary Health Centre
        elif FType == 'HWC-PHC/PHC':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 305th
            val_Description = [f"Cond{i}" for i in range(81)]
            # val_Description = [
            #                     'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC', 
            #                     'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC', 
            #                     'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC',
            #                     'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
            #                         'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
            #                         'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC',
            #                         'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
            #                         'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC', 
            #                         'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
            #                         'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
            #                             'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
            #                             'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
            #                                 'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
            #                                 'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
            #                                 'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
            #                                 'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                                 'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
            #                                     'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
            #                                     'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
            #                                     'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                     'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                     'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
            #                                     'Immunisation sessions held <=Immunisation sessions planned',
            #                                     'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
            #                                     'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis',
            #                                     'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female']



        # For Sub District Hospital
        elif FType == 'Community Health Centre':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 305th
            val_Description = val_Description = [f"Cond{i}" for i in range(84)]
            # val_Description = [
            #                     'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
            #                     'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
            #                     'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
            #                     'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
            #                     'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
            #                         'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
            #                         'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
            #                         'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
            #                         'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
            #                         'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
            #                             'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                             'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                             'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
            #                             'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
            #                             'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                                 'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
            #                                 'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
            #                                 'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
            #                                 'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                 'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
            #                                 'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                 'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                 'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
            #                                  'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC']
                                                        

        # For Primary Health Centre
        elif FType == 'Sub District Hospital':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 321th
            val_Description = val_Description = [f"Cond{i}" for i in range(87)]
            # val_Description = [
            #                     'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
            #                     'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
            #                     'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
            #                         'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
            #                         'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
            #                         'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
            #                             'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
            #                             'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
            #                             'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
            #                             'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
            #                                 'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                                 'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                                 'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
            #                                 'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
            #                                 'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                                 'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
            #                                     'Complications following male sterilization <= Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
            #                                     'Complications following female sterilization <= Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Postpartum sterilizations (within 7 days of delivery by minilap or concurrent with caesarean section) conducted + Number of Postpartum (within 48 hours of delivery) IUCD insertions',
            #                                     'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
            #                                         'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                         'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                         'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC']
            
        # For District Hospital
        elif FType == 'District Hospital':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 326th
            val_Description = val_Description = [f"Cond{i}" for i in range(89)]
            # val_Description = [
                                
            #                     'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
            #                     'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
            #                     'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
            #                     'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
            #                     'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    
            #                         'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
            #                         'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    
            #                         'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
            #                         'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                   
            #                             'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                       
            #                             'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                             'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
            #                             'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
            #                                 'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
            #                                 'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                           
            #                                 'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
            #                                 'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
            #                                 'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                           
            #                                 'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
            #                                 'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                 'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
            #                                 'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC']


    ###! TO GENERATE REPORT BY BLANKS
    ###! ----------------------------
    ###! ----------------------------

    def summaryReportByBlanks(self, df):
        global final_result_summ1, final_result_summ2, final_result_summ2_For_SubPlots,col_sum, dft_ARFacilityWise, dft_ARCheckWiseInc, \
                FList1, dft_ARCheckWiseINCBlnk, FList2, dft_FacilityWiseINCBlnk, dataframeForSheet6, dataframeForSheet7, FList3_copy, FList3, \
                    FList4, FList4_duplicate, dataframeForSheet4, dataframeForSheet5, FList2_copy, FList4_copy, df_Test_Plot_Inc_State, \
                        df_Test_Plot_INCBlnk_State, count_df_Test_Plot_Inc_State, count_df_Test_Plot_INCBlnk_State

        self.indicator_Description(df)

        '''
        ## First Summary Report
        ## ---------------------
        '''
        count_Inconsistent = []
        count_IncDueToBlank = []

        Columns = list(df_SummReport.columns.values.tolist())

        for col_name in Columns:

            c2 = df_SummReport[col_name].str.match("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum()
            count_Inconsistent.append(c2)

            c4 = df_SummReport[col_name].str.match("^Inconsistent due to blank\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum()
            count_IncDueToBlank.append(c4)

        ###! PRINTING length of ...
        ###? ************************************************************************************************************
        print(len(val_Description), len(count_Inconsistent), len(count_IncDueToBlank))   

        # To show facilities in a column
        colInterest = df['Facility Name'].to_numpy() # Give column names in addition

        #!''' For Inconsistent '''
        inconsistent_list = []

        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]
                
                pattern = re.compile("^Inconsistent\+[0-9a-zA-Z]*\.[0-9a-zA-Z]*")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])
                else:
                    continue

            inconsistent_list.append(temp)           

        #!''' For INCBlnk '''
        INCBlnk_list = []
        # TO show facility in the sheet5
        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]

                pattern = re.compile("^Inconsistent due to blank\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])
                else:
                    continue

            INCBlnk_list.append(temp)

        final_result_summ1 = pd.DataFrame({ "Conditions": df_SummReport.columns, 
                                                "Description": val_Description,
                                                    "Facilities(Name) Showing Inconsistent": inconsistent_list,
                                                        "Inconsistent": count_Inconsistent,
                                                            "Facilities (Name) Showing Inconsistent due to blank": INCBlnk_list,
                                                                "Inconsistent due to blank": count_IncDueToBlank,
                                                            })


        final_result_summ1 = final_result_summ1.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ1 = final_result_summ1.reset_index(drop=True)

        dataframeForSheet4 = final_result_summ1[['Conditions', 'Description', 'Inconsistent', 'Facilities(Name) Showing Inconsistent']]
        dataframeForSheet5 = final_result_summ1[['Conditions', 'Description', 'Inconsistent due to blank', 'Facilities (Name) Showing Inconsistent due to blank']]

        dataframeForSheet5 = dataframeForSheet5[dataframeForSheet5['Inconsistent due to blank']  !=  0]

        FList1 = final_result_summ1["Facilities(Name) Showing Inconsistent"].tolist()
        FList2 = final_result_summ1["Facilities (Name) Showing Inconsistent due to blank"].tolist()

        # Total number of rows in the upoaded dataset
        count_rows = df_SummReport.shape[0]
        
        FList2_copy = dataframeForSheet5["Facilities (Name) Showing Inconsistent due to blank"].tolist()    

        # Percentage for Validation Summary Sheet to show color codes
        final_result_summ1['PerIncSheet1'] = final_result_summ1['Inconsistent']/count_rows*100
        final_result_summ1['Inconsistent'].value_counts()
        final_result_summ1['PerINCBlnkrrSheet1'] = final_result_summ1['Inconsistent due to blank']/count_rows*100

        len0 = len(final_result_summ1['PerIncSheet1'])
        len1 = len(final_result_summ1['PerINCBlnkrrSheet1'])

        # Deleting unnecessary columns
        # del final_result_summ1['PerIncSheet1']
        # del final_result_summ1['PerINCBlnkrrSheet1']
            
        def select_col_SumSheet(X):
            # COLORS
            c = [   'background-color:  #EF5350',                  #   >=25% RED
                        'background-color: #FFAF00',            #   10 - 25% LIGHTER RED
                            'background-color: #C0C000',        #   5 - 10% MORE LIGHTER RED
                                'background-color: #00AF5F',    #   < 5% LIGHTEST RED
                                    ' ']

            mask_30 = (X['Inconsistent'] == 0)
            cnt30 = mask_30.values.sum()
            mask_29 = (X['Inconsistent due to blank'] == 0)
            cnt29 = mask_29.values.sum()

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[4], X.index, columns=X.columns)
            df1.loc[mask_29, 'Inconsistent due to blank'] = c[3]
            df1.loc[mask_30, 'Inconsistent'] = c[3]

            return df1

        # Remoiving inconsistent facility names from validation summary sheet
        final_result_summ1.drop(['Facilities(Name) Showing Inconsistent'], axis = 1, inplace=True)
        # Remoiving INCBlnk facility names from validation summary sheet
        final_result_summ1.drop(['Facilities (Name) Showing Inconsistent due to blank'], axis = 1, inplace=True)

        final_result_summ1 = final_result_summ1.style.apply(select_col_SumSheet, axis=None)


        '''
        ## Second Summary Report
        ## --------------------
        '''
        summ2_countInconsistent = []
        summ2_countProbableRErr = []
        All_Blank = []

        #? Iterating over indices of each row and calculating number of Blanks for each Facility Name 
        for index in range(len(df_SummReport)):
            '''   For no. of Inconsistent   '''
            if df_SummReport.iloc[index, :].str.match("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum() > 0:
                inconsistent = df_SummReport.iloc[index, :].str.match("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum()
                print('** Entered Yess-------- **')
                summ2_countInconsistent.append(inconsistent)
            else:
                print('** Entered else **')
                summ2_countInconsistent.append(0)
                
            
            '''   For no. of Inconsistent due to blank   '''
            if df_SummReport.iloc[index, :].str.match('^Probable Reporting Error by Blank [0-9]*.[0-9]*').sum() > 0:
                probableRErr = df_SummReport.iloc[index, :].str.match('^Probable Reporting Error by Blank [0-9]*.[0-9]*').sum()
                print('** Entered Yess-------- **')
                summ2_countProbableRErr.append(probableRErr)
            else:
                print('** Entered else **')
                summ2_countProbableRErr.append(0)

            blank = df_SummReport.iloc[index, :].str.match("Blank").sum()
            if blank == len(df_SummReport.columns):
                All_Blank.append('Yes')
            else:
                All_Blank.append('No')


        #########################################################   
        #  Facility Specific Inconsistent (Sheet 6)     
        
        ''' For Inconsistent '''
        # **********************
        inc_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*")
                if pattern.match(str(primString)):
                    txt = primString.replace('nan', "NA")
                    temp.append(txt)

            inc_list.append(temp)

        ''' For INCBlnk '''
        # *************
        incblnk_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^Probable Reporting Error by Blank")
                if pattern.match(str(primString)):
                    txt = primString.replace('nan', "NA")
                    temp.append(txt)

            incblnk_list.append(temp)


        #?#############################################################################################
        #?#############################################################################################
        final_result_summ2 = pd.DataFrame({
                                            "DATE": df['DATE'].values.tolist(),
                                                "State": df['State'].values.tolist(),
                                                    "District": df['District Name'].values.tolist(),
                                                        "Sub-district": df['Sub-District Name'].values.tolist(),
                                                            # "Block": df['Block Name'].tolist(),
                                                                "Facility Name": colInterest,
                                                                    "Inconsistent": summ2_countInconsistent,
                                                                        "Probable Reporting Error": summ2_countProbableRErr,
                                                                            "All Blank": All_Blank,
                                                                                "Checks (Inconsistent)" : inc_list,
                                                                                    "Checks (INCBlnk)": incblnk_list
                                                                                     })


        # Sorting in alphabetical  order
        final_result_summ2 = final_result_summ2.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ2 = final_result_summ2.reset_index(drop=True)

    
        Counter(final_result_summ2['District'])
        Counter(final_result_summ2['Inconsistent'])
        Counter(final_result_summ2['Probable Reporting Error'])


        df_Test_Plot_Inc_State = final_result_summ2.groupby('State').apply(lambda x: x[x['Inconsistent'] > 0].count())['Inconsistent'].nlargest(5)
        ###? Dictionary having count of Inconsistents state wise
        count_df_Test_Plot_Inc_State = df_Test_Plot_Inc_State.to_dict()

        df_Test_Plot_INCBlnk_State = final_result_summ2.groupby('State').apply(lambda x: x[x['Probable Reporting Error'] > 0].count())['Probable Reporting Error'].nlargest(5)
        ###? Dictionary having count of INCBlnk state wise
        count_df_Test_Plot_INCBlnk_State = df_Test_Plot_INCBlnk_State.to_dict()

        ##!## Extract names of top five states
        FINAL_RESULT_2 = final_result_summ2.copy()
        LIST_Test_Plot_Inc_State = list(count_df_Test_Plot_Inc_State)
        final_result_summ2_For_SubPlots = FINAL_RESULT_2[FINAL_RESULT_2['State'].isin(LIST_Test_Plot_Inc_State)]


        FList3 = final_result_summ2["Checks (Inconsistent)"].tolist()
        FList4 = final_result_summ2["Checks (INCBlnk)"].tolist()

        FList4_duplicate = list(filter(None, FList4))

        ##################### Facility-wise Inconsistencies ########################
        dataframeForSheet6 = final_result_summ2[['DATE', 'Facility Name', 'Sub-district', 'District', 'State', 'Inconsistent', 'Checks (Inconsistent)']]
        dataframeForSheet6 = dataframeForSheet6[['DATE','Facility Name', 'Sub-district', 'District', 'State', 'Inconsistent', 'Checks (Inconsistent)']].fillna('Not Available in Dataset')
        
        ##################### Facility-wise Inconsistencies due to blank ########################
        dataframeForSheet7 = final_result_summ2[['DATE', 'Facility Name', 'Sub-district', 'District', 'State', 'Probable Reporting Error', 'Checks (INCBlnk)']]
        dataframeForSheet7 = dataframeForSheet7[dataframeForSheet7['Probable Reporting Error']  !=  0]
        FList4_copy = dataframeForSheet7["Checks (INCBlnk)"].tolist()
        dataframeForSheet7 = dataframeForSheet7[['DATE', 'Facility Name', 'Sub-district', 'District', 'State', 'Probable Reporting Error', 'Checks (INCBlnk)']].fillna('Not Available in Dataset')


        '''  To find percentage Facility Type Wise  '''
                                                    
        # For Health Sub Centre
        if FType == 'HWC-SC/SC':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/11 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Probable Reporting Error']/ 11 * 100
        
        # For Primary Health Centre
        elif FType == 'HWC-PHC/PHC':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/26 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Probable Reporting Error']/ 26 * 100

        # For Community Health Centre
        elif FType == 'Community Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/24 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Probable Reporting Error']/ 24 * 100

        # For Sub District Hospital
        elif FType == 'Sub District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/22 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Probable Reporting Error']/ 22 * 100

        # For District Hospital
        elif FType == 'District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/22 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Probable Reporting Error']/ 22 * 100
        
        ## Deleting unnecessary columns
        del final_result_summ2['Checks (Inconsistent)']
        del final_result_summ2['Checks (INCBlnk)']

        try:
            del final_result_summ2['Sub-district']
        except:
            pass

        # del final_result_summ2['Block']

        def select_col(X):
            global c
            # COLORS
            # ******
            c = ['background-color:  #EF5350',                  #>=50% RED
                    'background-color: #FFAF00',                #25-50% ORANGE
                        'background-color: #C0C000',            #10-25% YELLOW
                            'background-color: #00FF00',        #5-10% L GREEN
                                'background-color: #00AF5F',    #<5% GREEN
                                    ' ']

                
            mask_AllBlank = (X['All Blank'] == 'Yes')   
            mask_16 = (X['Inconsistent'] == 0)
            mask_15 = (X['Probable Reporting Error'] == 0)

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[5], X.index, columns=X.columns)
            df1.loc[mask_AllBlank, 'All Blank'] = c[0]
            df1.loc[mask_15, 'Probable Reporting Error'] = c[4]
            df1.loc[mask_16, 'Inconsistent'] = c[4]
            return df1

        final_result_summ2 = final_result_summ2.style.apply(select_col, axis=None)

        return final_result_summ1, final_result_summ2, dataframeForSheet4, dataframeForSheet5, dataframeForSheet6, dataframeForSheet7



    ###! TO GENERATE REPORT BY CHECKS
    ###! ----------------------------
    ###! ----------------------------
    def summaryReportByChecks(self, df):
        global final_result_summ1, final_result_summ2, final_result_summ2_For_SubPlots,col_sum, dft_ARFacilityWise, dft_ARCheckWiseInc, \
                FList1, dft_ARCheckWiseINCBlnk, FList2, dft_FacilityWiseINCBlnkINCBlnk, dataframeForSheet6, dataframeForSheet7, FList3_copy, FList3, \
                    FList4, FList4_duplicate, dataframeForSheet4, dataframeForSheet5, FList2_copy, FList4_copy, df_Test_Plot_Inc_State, \
                        df_Test_Plot_INCBlnk_State, count_df_Test_Plot_Inc_State, count_df_Test_Plot_INCBlnk_State

        self.indicator_Description(df)

        '''
        ## First Summary Report
        ## ---------------------
        '''
        count_Inconsistent = []
        count_IncDueToBlank = []

        Columns = list(df_SummReport.columns.values.tolist())

        for col_name in Columns:
            
            try:
                c2 = df_SummReport[col_name].str.match("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum()
                count_Inconsistent.append(c2)

                c4 = df_SummReport[col_name].str.count("^Inconsistent due to blank\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum()
                count_IncDueToBlank.append(c4)
            except:
                count_Inconsistent.append(0)
                count_IncDueToBlank.append(0)
              

        print(len(val_Description), len(count_Inconsistent), len(count_IncDueToBlank))   

        # To show facilities in a column
        colInterest = df['Facility Name'].to_numpy() # Give column names in addition

        #!''' For Inconsistent '''
        inconsistent_list = []

        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]
                
                pattern = re.compile("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])
                else:
                    continue

            inconsistent_list.append(temp)           

        #!''' For INCBlnk '''
        INCBlnk_list = []
        # TO show facility in the sheet5
        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]

                pattern = re.compile("^Inconsistent due to blank\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])
                else:
                    continue

            INCBlnk_list.append(temp)      


        final_result_summ1 = pd.DataFrame({"Conditions": df_SummReport.columns, 
                                            "Description": val_Description,
                                                "Facilities(Name) Showing Inconsistent": inconsistent_list,
                                                    "Inconsistent": count_Inconsistent,
                                                        "Facilities (Name) Showing Inconsistent due to blank": INCBlnk_list,
                                                            "Inconsistent due to blank": count_IncDueToBlank,
                                                            })

        final_result_summ1 = final_result_summ1.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ1 = final_result_summ1.reset_index(drop=True)

        dataframeForSheet4 = final_result_summ1[['Conditions', 'Description', 'Inconsistent', 'Facilities(Name) Showing Inconsistent']]
        dataframeForSheet5 = final_result_summ1[['Conditions', 'Description', 'Inconsistent due to blank', 'Facilities (Name) Showing Inconsistent due to blank']]

        dataframeForSheet5 = dataframeForSheet5[dataframeForSheet5['Inconsistent due to blank']  !=  0]

        FList1 = final_result_summ1["Facilities(Name) Showing Inconsistent"].tolist()
        FList2 = final_result_summ1["Facilities (Name) Showing Inconsistent due to blank"].tolist()

        # Total number of rows in the upoaded dataset
        count_rows = df_SummReport.shape[0]
        
        FList2_copy = dataframeForSheet5["Facilities (Name) Showing Inconsistent due to blank"].tolist()    

        # Percentage for Validation Summary Sheet to show color codes
        final_result_summ1['PerIncSheet1'] = final_result_summ1['Inconsistent']/count_rows*100
        final_result_summ1['Inconsistent'].value_counts()
        final_result_summ1['PerINCBlnkrrSheet1'] = final_result_summ1['Inconsistent due to blank']/count_rows*100

        len0 = len(final_result_summ1['PerIncSheet1'])
        len1 = len(final_result_summ1['PerINCBlnkrrSheet1'])

        # Deleting unnecessary columns
        # del final_result_summ1['PerIncSheet1']
        # del final_result_summ1['PerINCBlnkrrSheet1']
            
        def select_col_SumSheet(X):
            # COLORS
            c = ['background-color:  #EF5350',                  #   >=25% RED
                        'background-color: #FFAF00',            #   10 - 25% LIGHTER RED
                            'background-color: #C0C000',        #   5 - 10% MORE LIGHTER RED
                                'background-color: #00AF5F',    #   < 5% LIGHTEST RED
                                    ' ']

            mask_30 = (X['Inconsistent'] == 0)
            cnt30 = mask_30.values.sum()
            mask_29 = (X['Inconsistent due to blank'] == 0)
            cnt29 = mask_29.values.sum()

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[4], X.index, columns=X.columns)
            df1.loc[mask_29, 'Inconsistent due to blank'] = c[3]
            df1.loc[mask_30, 'Inconsistent'] = c[3]

            return df1

        # Remoiving inconsistent facility names from validation summary sheet
        final_result_summ1.drop(['Facilities(Name) Showing Inconsistent'], axis = 1, inplace=True)
        # Remoiving INCBlnk facility names from validation summary sheet
        final_result_summ1.drop(['Facilities (Name) Showing Inconsistent due to blank'], axis = 1, inplace=True)

        final_result_summ1 = final_result_summ1.style.apply(select_col_SumSheet, axis=None)


        '''
        ## Second Summary Report
        ## --------------------
        '''
        summ2_countInconsistent = []
        summ2_countProbableRErr = []
        All_Blank = []

        # Iterating over indices of each row and calculating number of Blanks for each Facility Name 
        for index in range(len(df_SummReport)):
            '''   For no. of Inconsistent   '''
            if df_SummReport.iloc[index, :].str.match("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum() > 0:
                inconsistent = df_SummReport.iloc[index, :].str.match("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*").sum()
                summ2_countInconsistent.append(inconsistent)
            else:
                summ2_countInconsistent.append(0)
            
            '''   For no. of Probable Reporting Errors   '''
            if df_SummReport.iloc[index, :].str.match('^Inconsistent due to blank\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*').sum() > 0:
                probableRErr = df_SummReport.iloc[index, :].str.match('^Inconsistent due to blank\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*').sum()
                summ2_countProbableRErr.append(probableRErr)
            else:
                summ2_countProbableRErr.append(0)

            blank = df_SummReport.iloc[index, :].str.count("Blank").sum()
            if blank == len(df_SummReport.columns):
                All_Blank.append('Yes')
            else:
                All_Blank.append('No')


        #########################################################   
        #  Facility Specific Inconsistent (Sheet 6)     
        
        ''' For Inconsistent '''
        # **********************
        inc_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^Inconsistent\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*")
                if pattern.match(str(primString)):
                    txt = primString.replace('nan', "NA")
                    temp.append(txt)

            inc_list.append(temp)

        ''' For INCBlnk '''
        # *************
        incblnk_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^Inconsistent due to blank\s+[0-9a-zA-Z]*\.[0-9a-zA-Z]*")
                if pattern.match(str(primString)):
                    txt = primString.replace('nan', "NA")
                    temp.append(txt)

            incblnk_list.append(temp)


        ###################################################
        final_result_summ2 = pd.DataFrame({
                                            "DATE": df['DATE'].values.tolist(),
                                                "State": df['State'].values.tolist(),
                                                    "District": df['District Name'].values.tolist(),
                                                        "Sub-district": df['Sub-District Name'].values.tolist(),
                                                            # "Block": df['Block Name'].tolist(),
                                                                "Facility Name": colInterest,
                                                                    "Inconsistent": summ2_countInconsistent,
                                                                        "Inconsistent due to blank": summ2_countProbableRErr,
                                                                            "All Blank": All_Blank,
                                                                                "Checks (Inconsistent)" : inc_list,
                                                                                    "Checks (Inconsistent due to blank)": incblnk_list
                                                                                    })


        # Sorting in alphabetical  order
        final_result_summ2 = final_result_summ2.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ2 = final_result_summ2.reset_index(drop=True)

    
        Counter(final_result_summ2['District'])
        Counter(final_result_summ2['Inconsistent'])
        Counter(final_result_summ2['Inconsistent due to blank'])

        df_Test_Plot_Inc_State = final_result_summ2.groupby('State').apply(lambda x: x[x['Inconsistent'] > 0].count())['Inconsistent'].nlargest(5)
        ###? Dictionary having count of Inconsistents state wise
        count_df_Test_Plot_Inc_State = df_Test_Plot_Inc_State.to_dict()

        df_Test_Plot_INCBlnk_State = final_result_summ2.groupby('State').apply(lambda x: x[x['Inconsistent due to blank'] > 0].count())['Inconsistent due to blank'].nlargest(5)
        ###? Dictionary having count of INCBlnk state wise
        count_df_Test_Plot_INCBlnk_State = df_Test_Plot_INCBlnk_State.to_dict()

        ##!## Extract names of top five states
        FINAL_RESULT_2 = final_result_summ2.copy()
        LIST_Test_Plot_Inc_State = list(count_df_Test_Plot_Inc_State)
        final_result_summ2_For_SubPlots = FINAL_RESULT_2[FINAL_RESULT_2['State'].isin(LIST_Test_Plot_Inc_State)]

        FList3 = final_result_summ2["Checks (Inconsistent)"].tolist()
        FList4 = final_result_summ2["Checks (Inconsistent due to blank)"].tolist()

        FList4_duplicate = list(filter(None, FList4))

        ##################### Facility-wise Inconsistencies ########################
        dataframeForSheet6 = final_result_summ2[['DATE', 'Facility Name', 'Sub-district', 'District', 'State', 'Inconsistent', 'Checks (Inconsistent)']]
        dataframeForSheet6 = dataframeForSheet6[['DATE','Facility Name', 'Sub-district', 'District', 'State', 'Inconsistent', 'Checks (Inconsistent)']].fillna('Not Available in Dataset')
        
        ##################### Facility-wise Inconsistencies due to blank ########################
        dataframeForSheet7 = final_result_summ2[['DATE', 'Facility Name', 'Sub-district', 'District', 'State', 'Inconsistent due to blank', 'Checks (Inconsistent due to blank)']]
        dataframeForSheet7 = dataframeForSheet7[dataframeForSheet7['Inconsistent due to blank']  !=  0]
        FList4_copy = dataframeForSheet7["Checks (Inconsistent due to blank)"].tolist()
        dataframeForSheet7 = dataframeForSheet7[['DATE', 'Facility Name', 'Sub-district', 'District', 'State', 'Inconsistent due to blank', 'Checks (Inconsistent due to blank)']].fillna('Not Available in Dataset')


        '''  To find percentage Facility Type Wise   '''
        ###? -------------------------------------------
                                                    
        # For Health Sub Centre
        if FType == 'HWC-SC/SC':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/11 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Inconsistent due to blank']/ 11 * 100
        
        # For Primary Health Centre
        elif FType == 'HWC-PHC/PHC':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/26 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Inconsistent due to blank']/ 26 * 100

        # For Community Health Centre
        elif FType == 'Community Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/24 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Inconsistent due to blank']/ 24 * 100

        # For Sub District Hospital
        elif FType == 'Sub District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/22 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Inconsistent due to blank']/ 22 * 100

        # For District Hospital
        elif FType == 'District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/22 * 100
            final_result_summ2['PercentageINCBlnkrr'] = final_result_summ2['Inconsistent due to blank']/ 22 * 100
        
        ## Deleting unnecessary columns
        del final_result_summ2['Checks (Inconsistent)']
        del final_result_summ2['Checks (Inconsistent due to blank)']

        try:
            del final_result_summ2['Sub-district']
        except:
            pass


        def select_col(X):
            global c
            # COLORS
            # ******
            c = ['background-color:  #EF5350',                  #>=50% RED
                    'background-color: #FFAF00',                #25-50% ORANGE
                        'background-color: #C0C000',            #10-25% YELLOW
                            'background-color: #00FF00',        #5-10% L GREEN
                                'background-color: #00AF5F',    #<5% GREEN
                                    ' ']

                
            mask_AllBlank = (X['All Blank'] == 'Yes')   
            mask_16 = (X['Inconsistent'] == 0)
            mask_15 = (X['Inconsistent due to blank'] == 0)

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[5], X.index, columns=X.columns)
            df1.loc[mask_AllBlank, 'All Blank'] = c[0]
            df1.loc[mask_15, 'Inconsistent due to blank'] = c[4]
            df1.loc[mask_16, 'Inconsistent'] = c[4]
            return df1

        final_result_summ2 = final_result_summ2.style.apply(select_col, axis=None)

        return final_result_summ1, final_result_summ2, dataframeForSheet4, dataframeForSheet5, dataframeForSheet6, dataframeForSheet7


    ####? ----------------------------------------- EXPORT FILE ------------------------------------
    ####* ==========================================================================================
    def export(self):
        global filename, table_result1, table_result2, table_result3, table_result4, table_result5, table_result6
     
        # Save file dialog
        UI_Val = Ui_Dialog_Validate()
        UI_Val.exec_()
        
        # try:
        if UI_Val.Flag == 'Summary_Report_By_Checks':
            table_result1, table_result2, table_result3, table_result4, table_result5, table_result6  = self.summaryReportByChecks(df)
            filename = QFileDialog.getSaveFileName(TabWidget, "Save to Excel", "Validated Results Summary Sheet",
                                            "Excel Spreadsheet (*.xlsx);;"
                                            "All Files (*)")[0]
        elif UI_Val.Flag == 'Summary_Report_By_Blanks':
            table_result1, table_result2, table_result3, table_result4, table_result5, table_result6  = self.summaryReportByBlanks(df)
            filename = QFileDialog.getSaveFileName(TabWidget, "Save to Excel", "Validated Results Summary Sheet",
                                            "Excel Spreadsheet (*.xlsx);;"
                                            "All Files (*)")[0]
                
        # except:
        # try:
        # if UI_Val.Flag == 'Summary_Report_By_Blanks':
        #     table_result1, table_result2, table_result3, table_result4, table_result5, table_result6  = self.summaryReportByBlanks(df)
        #     filename = QFileDialog.getSaveFileName(TabWidget, "Save to Excel", "Validated Results Summary Sheet",
        #                                     "Excel Spreadsheet (*.xlsx);;"
        #                                     "All Files (*)")[0]
        # elif UI_Val.Flag == 'Summary_Report_By_Checks':
        #     table_result1, table_result2, table_result3, table_result4, table_result5, table_result6  = self.summaryReportByChecks(df)
        #     filename = QFileDialog.getSaveFileName(TabWidget, "Save to Excel", "Validated Results Summary Sheet",
        #                                     "Excel Spreadsheet (*.xlsx);;"
        #                                     "All Files (*)")[0]
        # except:
        #     pass
        # finally:
        #     pass
        
        try:
            if table_result1 != "":
                self.saveFile()
        except:
            pass

        try:
            return df, filename, table_result1, table_result2, table_result3, table_result4, table_result5, table_result6
        except:
            pass


    def saveFile(self):

        new_list = [["Description about the output sheets:"]]
        table_result_content = pd.DataFrame(new_list)

        if filename != "":
            self.reopenAndDesignExcel(filename, table_result_content, table_result1, table_result2, table_result3, table_result4, table_result5, table_result6)


    ###! Saving Validated Data
    ###  ---------------------
    def saveValidatedData(self, *args, **kwargs):

        try:
            # Save file dialog
            filename = QFileDialog.getSaveFileName(TabWidget, "Save to Excel", "Validated Data Summary Sheet",
                                                        "Excel Spreadsheet (*.xlsx);;"
                                                            "All Files (*)")[0]

            
            writer = pd.ExcelWriter(filename, engine='xlsxwriter')

            # Convert the dataframe to an XlsxWriter Excel object.
            df.to_excel(writer, startrow = 2, sheet_name='validated_data')

            workbook  = writer.book
            worksheet = writer.sheets['validated_data']

            text = str(new_header[0:1])
            
            text = text.replace("State Code","")
            text = text.replace("Name: 0, dtype: object", "")
            text = text.replace("\n"," ")
            worksheet.write(0, 0, text)

            header_format = workbook.add_format({
                'bold': True,
                'border': False})


            # Close the Pandas Excel writer and output the Excel file.
            writer.save()

            # self.popup.close()

            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("\n Validated Data File Exported / Validated Data फ़ाइल निर्यात की गई ")
            # Set the title of the window
            self.msg.setWindowTitle(" Data File is exported / डेटा फ़ाइल निर्यात की गई ")
            # Display the message box
            self.msg.show()

        except:
            msg = QMessageBox()
            msg.setWindowTitle("Saving File Error Message / फ़ाइल सहेजें त्रुटि संदेश")
            msg.setIcon(QMessageBox.Critical)
            msg.setText("\n First validate your uploaded data ! / पहले अपने अपलोड किए गए डेटा को सत्यापित करें !")
            msg.exec()
        finally:
            pass


    ###################################################################################
    ########################## Reopen excel and start processes
    def reopenAndDesignExcel(self, filename, table_result_content, table_result1, table_result2, table_result3, table_result4, table_result5, table_result6):
        
        table_result3 = table_result3[table_result3.Inconsistent  !=  0]
        table_result5 = table_result5[table_result5.Inconsistent  !=  0]
        # table_result4 = table_result3[table_result3.Inconsistent  !=  0]
        # table_result6 = table_result5[table_result5.Inconsistent  !=  0]
        
        # Taking transpose of data 
        table_result3 = table_result3[:101].T
        table_result4 = table_result4[:101].T
        table_result5 = table_result5[:101].T
        table_result6 = table_result6[:101].T

        # try:
        # exporting to excel
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer: 
            table_result_content.to_excel(writer, sheet_name='Description')
            table_result2.to_excel(writer, sheet_name='Facility-level summary')
            table_result5.to_excel(writer, sheet_name='Facility-wise inconsistencies')
            table_result6.to_excel(writer, sheet_name='Facility-wise Inc due to blank')
            table_result1.to_excel(writer, sheet_name='Validation rule wise summary')
            table_result3.to_excel(writer, sheet_name='Inconsistency-wise facilities')
            table_result4.to_excel(writer, sheet_name='Inc due to blank-wise facility')
        

        # PALETTES
        workbook = load_workbook(filename)
        sheet_0 = workbook['Description']
        sheet = workbook['Facility-level summary']
        sheet_4 = workbook['Facility-wise inconsistencies']
        sheet_5 = workbook['Facility-wise Inc due to blank']
        sheet_1 = workbook['Validation rule wise summary']
        sheet_2 = workbook['Inconsistency-wise facilities']
        sheet_3 = workbook['Inc due to blank-wise facility']
        

        # Activating sheets 
        workbook.active = sheet_0
        workbook.active = sheet
        workbook.active = sheet_1
        workbook.active = sheet_2
        workbook.active = sheet_3
        workbook.active = sheet_4
        workbook.active = sheet_5

        # ### Adding Gridlines
        sheet.sheet_view.showGridLines = False
        sheet_1.sheet_view.showGridLines = False

        '''
        ADDING THICK BORDERS
        '''
        from openpyxl.styles.borders import Border, Side
        thick_border = Border(left=Side(style='thick'), 
                            right=Side(style='thick'), 
                            top=Side(style='thick'), 
                            bottom=Side(style='thick'))



        '''
        *********************************** sheet_0 :: Description ***********************************
        '''
        
        ## Bordering Description Sheet
        for i in range(4, 12):
            sheet_0.cell(row=i+1, column=2).border = thick_border
            sheet_0.cell(row=i+1, column=3).border = thick_border

        for i in range(16, 20):
            sheet_0.cell(row=i+1, column=2).border = thick_border
            sheet_0.cell(row=i+1, column=3).border = thick_border

        sheet_0.sheet_view.showGridLines = False
        sheet_0.merge_cells('A1:B2')
        cell_A1 = sheet_0.cell(row= 1, column= 1)
        cell_A1.value = 'Description about the output sheets:'  
        cell_A1.alignment = Alignment(horizontal='center', vertical='center', indent=0) 
        cell_A1.fill = PatternFill("solid", fgColor="00003366")
        cell_A1.font = Font(color="00FFFFFF", size = 14, bold = True)

        sheet_0.column_dimensions['A'].width = 15
        sheet_0.column_dimensions['B'].width = 30
        sheet_0.column_dimensions['C'].width = 80
        sheet_0.column_dimensions['D'].width = 80

        sheet_0['A3'] = "Facility Type: "
        sheet_0['A4'] = "Month, Year: "

        sheet_0['B3'] = self.lineEdit_2.text()
        sheet_0['B4'] = items_date

        sheet_0.row_dimensions[3].height = 25
        sheet_0.row_dimensions[4].height = 15

        sheet_0.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center') 
        sheet_0.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center') 
        sheet_0.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center') 
        sheet_0.cell(row=4, column=2).alignment = Alignment(horizontal='center', vertical='center')

        sheet_0.cell(row=3, column=3).alignment = Alignment(horizontal='right')
        sheet_0['B5'] = "Sheet Names"
        sheet_0['B6'] = "Description"
        sheet_0['B7'] = "Facility-level summary"
        sheet_0['B8'] = "Check with Inconistencies"
        sheet_0['B9'] = "Facility-wise Inconsistencies due to blank"
        sheet_0['B10'] = "Validation rule wise summary"
        sheet_0['B11'] = "Facility with Inconsistencies"
        sheet_0['B12'] = "Facility with Inconsistencies due to blank"

        sheet_0['B16'] = "Validation Check Outcome Definition : "
        sheet_0['B17'] = "Consistent"
        sheet_0['B18'] = "Inconsistent"
        sheet_0['B19'] = "Inconsistent due to blank"
        sheet_0['B20'] = "Blank"

        sheet_0['C4'] = ""
        sheet_0['C5'] = "Details"
        sheet_0['C6'] = "Description of sheets, important terminologies and other explanations"
        sheet_0['C7'] = "This sheet gives the counts of errors corresponding to each facility name. The colour coding is done as per the buckets, considering the percentage of the number of inconsistent/inconsistent due to balnk out of the total validation checks in that facility type for each facility name. This is also shown graphically below the bucketing table. Clicking on the figures of the original table, the user will be directed to the “Checks giving inconsistent” or “Checks giving inconsistent due to blank” tabs, showing the validation checks for which, the errors creeped in."
        sheet_0['C8'] = "This sheet provides list of inconsistent checks for each facility within dataset highlighting the values of data items involved in the check"
        sheet_0['C9'] = "This sheet provides list of validation checks with 'inconsistent due to blank outcome' for each facility within dataset highlighting the values of data items involved in the check"
        sheet_0['C10'] = 'This sheet This sheet gives us the count of the facilities giving "inconsistent" and "inconsistent due to blank" error within the datasets for each validation check. The colour coding is done as per the buckets, which were created considering the percentage of the number of facility names out of all the facilities who have reported inconsistent/PRE. This is also shown graphically below the bucketing table. Clicking on the figures of the original table, the user will be directed to the “Facility with inconsistent” or “Facility with PRE” tabs, showing the facilities for which, the errors creeped in.'
        sheet_0['C11'] = "This sheet provides list of facilites depicting inconsistent error for each validation check highlighting the description of check"
        sheet_0['C12'] = "This sheet provides list of facilites depicting inconsistent due to blank error for each validation check highlighting the description of check"
        
        sheet_0['C16'] = ""
        sheet_0['C17'] = "The validation check holds true and needs no scrutiny."
        sheet_0['C18'] = "The validation check fails and the inconsistent data item is flagged. "
        sheet_0['C19'] = "Data items on one side of validation check are blank and the inconsistent due to blank data item is flagged"
        sheet_0['C20'] = " Data items on both the sides of validation check are blank."

        sheet_0['B16'].font = Font(size = 14, bold = True)
        sheet_0['B16'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['C16'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        # sheet_0['C4'].fill = PatternFill(fgColor="FFFFCC", fill_type = "solid")
        sheet_0['B5'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['B5'].font = Font(bold = True)
        sheet_0['C5'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['C5'].font = Font(bold = True)

        for i in range(16, 20):
            sheet_0.cell(row=i+1, column=2).border = thick_border
            sheet_0.cell(row=i+1, column=3).border = thick_border

        sheet_0['B23'] = "Examples"
        sheet_0['B23'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['B23'].font = Font(size = 14, bold = True)
        sheet_0['C23'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")

        sheet_0.merge_cells('B25:B29')  
  
        cell_B25 = sheet_0.cell(row= 25, column= 2)  
        cell_B25.value = '1.1.1 Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)'  
        cell_B25.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_B25.fill = PatternFill("solid", fgColor="00003366")
        cell_B25.font = Font(color="00FFFFFF")
        

        sheet_0.merge_cells('C25:C29')  
  
        cell_C25 = sheet_0.cell(row= 25, column= 3)  
        cell_C25.value = '1.1 Total number of pregnant women registered for ANC'  
        cell_C25.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_C25.fill = PatternFill("solid", fgColor="00003366")
        cell_C25.font = Font(color="00FFFFFF")
        

        sheet_0.merge_cells('D25:D29')  
  
        cell_D25 = sheet_0.cell(row= 25, column= 4)  
        cell_D25.value = 'Outcome (for 1.1.1 <= 1.1)'  
        cell_D25.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_D25.fill = PatternFill("solid", fgColor="00003366")
        cell_D25.font = Font(color="00FFFFFF")
        
        
        sheet_0['B30'] = 'Value'
        sheet_0['B30'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B30'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['B30'].font = Font(color="00000000")
        sheet_0['B30'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))
        

        sheet_0['B31'] = 'Value'
        sheet_0['B31'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B31'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['B31'].font = Font(color="00000000")
        sheet_0['B31'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B32'] = 'Null'
        sheet_0['B32'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B32'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['B32'].font = Font(color="00000000")
        sheet_0['B32'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B33'] = 'Value'
        sheet_0['B33'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B33'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['B33'].font = Font(color="00000000")
        sheet_0['B33'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B34'] = 'Blank'
        sheet_0['B34'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B34'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['B34'].font = Font(color="00000000")
        sheet_0['B34'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C30'] = 'Value'
        sheet_0['C30'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C30'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C30'].font = Font(color="00000000", bold = False)
        sheet_0['C30'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C31'] = 'Value'
        sheet_0['C31'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C31'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C31'].font = Font(color="00000000", bold = False)
        sheet_0['C31'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C32'] = 'Value'
        sheet_0['C32'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C32'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C32'].font = Font(color="00000000", bold = False)
        sheet_0['C32'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C33'] = 'Null'
        sheet_0['C33'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C33'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C33'].font = Font(color="00000000", bold = False)
        sheet_0['C33'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C34'] = 'Blank'
        sheet_0['C34'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C34'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C34'].font = Font(color="00000000", bold = False)
        sheet_0['C34'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D30'] = 'Consistent'
        sheet_0['D30'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D30'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D30'].font = Font(color="00000000", bold = False)
        sheet_0['D30'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D31'] = 'Inconsistent (when condition fails)'
        sheet_0['D31'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D31'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D31'].font = Font(color="00000000", bold = False)
        sheet_0['D31'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D32'] = 'Inconsistent due to blank'
        sheet_0['D32'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D32'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D32'].font = Font(color="00000000", bold = False)
        sheet_0['D32'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D33'] = 'Inconsistent due to blank'
        sheet_0['D33'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D33'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D33'].font = Font(color="00000000", bold = False)
        sheet_0['D33'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D34'] = 'Blank'
        sheet_0['D34'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D34'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D34'].font = Font(color="00000000", bold = False)
        sheet_0['D34'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))


        sheet_0.row_dimensions[30].height = 15
        sheet_0.row_dimensions[31].height = 15
        sheet_0.row_dimensions[32].height = 15
        sheet_0.row_dimensions[33].height = 15
        sheet_0.row_dimensions[34].height = 15


        ### Second
        sheet_0['B37'] = "Recurring data items [service for one data items may be provided over months]:  Considering permissible limit of +-50% i.e., if the disparity in the two data items is more than the limit then it is a probable reporting error."
        sheet_0['B37'].font = Font(size = 12, bold = True)
        
        sheet_0.merge_cells('B38:B42')  
  
        cell_B37 = sheet_0.cell(row= 38, column= 2)  
        cell_B37.value = '(1.1.2)  Total ANC footfall/cases (Old cases + New Registration) attended '  
        cell_B37.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_B37.fill = PatternFill("solid", fgColor="00003366")
        cell_B37.font = Font(color="00FFFFFF")

        sheet_0.merge_cells('C38:C42')  
  
        cell_C50 = sheet_0.cell(row= 38, column= 3)  
        cell_C50.value = '(1.2.1) Number of PW given Td1 (Tetanus Diptheria dose 1) + (1.2.2) Number of PW given Td2 (Tetanus Diptheria dose 2) + (1.2.3) Number of PW given Td Booster (Tetanus Diptheria dose booster)'  
        cell_C50.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_C50.fill = PatternFill("solid", fgColor="00003366")
        cell_C50.font = Font(color="00FFFFFF")

        sheet_0.merge_cells('D38:D42')  
  
        cell_D50 = sheet_0.cell(row= 38, column= 4)  
        cell_D50.value = 'Outcome (for 1.1.2 >= 1.2.1 + 1.2.2 + 1.2.3)'  
        cell_D50.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_D50.fill = PatternFill("solid", fgColor="00003366")
        cell_D50.font = Font(color="00FFFFFF")

        sheet_0['B37'].fill = PatternFill("solid", fgColor="FFFF00")

        sheet_0['B43'] = 'Value'
        sheet_0['B43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B43'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['B43'].font = Font(color="00000000")
        sheet_0['B43'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B44'] = 'Value'
        sheet_0['B44'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B44'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['B44'].font = Font(color="00000000")
        sheet_0['B44'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B45'] = 'Null'
        sheet_0['B45'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B45'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['B45'].font = Font(color="00000000")
        sheet_0['B45'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B46'] = 'Value'
        sheet_0['B46'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B46'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['B46'].font = Font(color="00000000")
        sheet_0['B46'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B47'] = 'Blank'
        sheet_0['B47'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B47'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['B47'].font = Font(color="00000000")
        sheet_0['B47'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C43'] = 'Value + value + value'
        sheet_0['C43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C43'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C43'].font = Font(color="00000000", bold = False)
        sheet_0['C43'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C44'] = 'Value + value + value'
        sheet_0['C44'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C44'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C44'].font = Font(color="00000000", bold = False)
        sheet_0['C44'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C45'] = 'Value + value + value'
        sheet_0['C45'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C45'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C45'].font = Font(color="00000000", bold = False)
        sheet_0['C45'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C46'] = 'Null + Null + Null'
        sheet_0['C46'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C46'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C46'].font = Font(color="00000000", bold = False)
        sheet_0['C46'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C47'] = 'Blank + Blank + Blank'
        sheet_0['C47'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C47'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C47'].font = Font(color="00000000", bold = False)
        sheet_0['C47'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D43'] = 'Consistent'
        sheet_0['D43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D43'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D43'].font = Font(color="00000000", bold = False)
        sheet_0['D43'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D44'] = 'Inconsistent (when condition fails)'
        sheet_0['D44'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D44'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D44'].font = Font(color="00000000", bold = False)
        sheet_0['D44'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D45'] = 'Inconsistent due to blank'
        sheet_0['D45'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D45'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D45'].font = Font(color="00000000", bold = False)
        sheet_0['D45'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D46'] = 'Inconsistent due to blank'
        sheet_0['D46'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D46'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D46'].font = Font(color="00000000", bold = False)
        sheet_0['D46'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D47'] = 'Blank'
        sheet_0['D47'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D47'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D47'].font = Font(color="00000000", bold = False)
        sheet_0['D47'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0.row_dimensions[43].height = 15
        sheet_0.row_dimensions[44].height = 15
        sheet_0.row_dimensions[45].height = 15
        sheet_0.row_dimensions[46].height = 15
        sheet_0.row_dimensions[47].height = 15

        ### WRAPPING TEXT IN DESCRIPTION SHEET
        for rows in sheet_0.iter_rows():
            for cell in rows:
                cell.alignment = Alignment(wrapText=True)

        workbook.save(filename=filename)

        '''
        **************************************************************************************************
        '''

        '''
        *********************************** sheet :: Facility Level Summary ***********************************
        '''

        ''' THE CODE FOR GENERATING HYPERLINKS '''
        # Function to generate sequences according to the excel sheet
        def excel_cols_link3():
            n1 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n1))
                n1 += 1

        # Function to generate sequences according to the excel sheet
        def excel_cols_link4():
            n2 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n2))
                n2 += 1

        ## Formatting Sheet{Facility Level Summary}
        try:
            sheet.move_range("A1:J10000", rows=1)
        except:
            sheet.move_range("A1:J25000", rows=1)
            try:
                sheet.move_range("A1:J50000", rows=1)
            except:
                sheet.move_range("A1:J100000", rows=1)
            finally:
                sheet.move_range("A1:J1048576", rows=1)


        ## ADDING HEADER IN Facility Level Summary
        sheet.oddHeader.center.text = "Facility Level Summary"
        sheet.oddHeader.center.size = 14
        sheet.oddHeader.center.font = "Tahoma,Bold"
        sheet.oddHeader.center.color = "CC3366"
        sheet.cell(row=1, column=4).value = 'Facility Level Summary'
        sheet.cell(row=2, column=1).value = 'Sr. No'
        sheet.cell(row=2, column=2).value = 'Date'
        sheet.cell(row=2, column=5).value = 'Facility Name'
        sheet['D1'].font = Font(size = 18, bold = True, color="003679") 

        ## EXPANDING ROWS OF SHEET (FACILITY LEVEL SUMMARY)
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 25

       
        '''
        !------------------------------------------------------------------
        #####! ADDING THE COLOR CODES TO THE SHEET (FACILITY LEVEL SUMMARY)
        !------------------------------------------------------------------
        '''
    
        cnt1, cnt2, cnt3, cnt4, cnt5 = 0, 0, 0, 0, 0
        for i in range(len(FList3)):

            ### Deleting unwanted rows
            if sheet.cell(row=i+3, column=9).value == None:
                sheet.delete_rows(i+3, 1)
            elif sheet.cell(row=i+3, column=9).value >= 50:
                cnt1 += 1
            elif sheet.cell(row=i+3, column=9).value < 50 and sheet.cell(row=i+3, column=9).value >= 25:
                cnt2 += 1
            elif sheet.cell(row=i+3, column=9).value < 25 and sheet.cell(row=i+3, column=9).value >= 10:
                cnt3 += 1
            elif sheet.cell(row=i+3, column=9).value < 10 and sheet.cell(row=i+3, column=9).value >= 5:
                cnt4 += 1
            elif sheet.cell(row=i+3, column=9).value < 5:
                cnt5 += 1

            for j in range(len(FList3[i])):
                '''
                # Create hyperlink to relevant cell
                '''
                link1 = "#'Facility-wise Inconsistencies'!B2"

                l1 = list(itertools.islice(excel_cols_link3(), len(FList3)+1))

                #update link
                link1 = link1.replace("B", l1[i+1])

                sheet.cell(row=i+3, column=6).hyperlink = link1
                sheet.cell(row=i+3, column=6).style = "Hyperlink"

                if sheet.cell(row=i+3, column=9).value >= 50:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='EF5350', fill_type = "solid") 
                elif sheet.cell(row=i+3, column=9).value < 50 and sheet.cell(row=i+3, column=9).value >= 25:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet.cell(row=i+3, column=9).value < 25 and sheet.cell(row=i+3, column=9).value >= 10:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet.cell(row=i+3, column=9).value < 10 and sheet.cell(row=i+3, column=9).value >= 5:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='00FF00', fill_type = "solid")
                elif sheet.cell(row=i+3, column=9).value < 5:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")


        link2 = "#'Facility-wise Inc due to blank'!B2"
        l2 = list(itertools.islice(excel_cols_link4(), dataframeForSheet7.shape[0]+1))

        # for i in range(len(FList4)):
        k = 0
        for i in range(len(FList4)):
            if sheet.cell(row=i+3, column=7).value == 0:
                k = k + 1 
            if sheet.cell(row=i+3, column=7).value != 0 and sheet.cell(row=i+2, column=7).value == 0:
                if k == 1:
                    link2 = "#'Facility-wise Inc due to blank'!"+ l2[i] +"2"
                    print('i=', i, "cell=", sheet.cell(row=i+3, column=7), "Value=", sheet.cell(row=i+3, column=7).value, 'link=', link2)
                elif k > 1:
                    link2 = "#'Facility-wise Inc due to blank'!"+ l2[i+1-k] +"2"
                    print('i=', i, "cell=", sheet.cell(row=i+3, column=7), "Value=", sheet.cell(row=i+3, column=7).value, 'link=', link2)

            elif sheet.cell(row=i+3, column=6).value != 0 and sheet.cell(row=i+2, column=7).value != 0:
                link2 = "#'Facility-wise Inc due to blank'!"+ l2[i+1-k] +"2"   
                print('i=', i, "celltype2=", sheet.cell(row=i+3, column=7), "Value=", sheet.cell(row=i+3, column=7).value, 'link=', link2)
            else:
                link2=None

            sheet.cell(row=i+3, column=7).hyperlink = link2
            sheet.cell(row=i+3, column=7).style = "Hyperlink"



        cnt6, cnt7, cnt8, cnt9, cnt10 = 0, 0, 0, 0, 0
        for i in range(len(FList4)):
            if sheet.cell(row=i+3, column=10).value == None:
                sheet.delete_rows(i+3, 1)
            elif sheet.cell(row=i+3, column=10).value >= 50:
                cnt6 += 1
                sheet.cell(row=i+3, column=7).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
            elif sheet.cell(row=i+3, column=10).value < 50 and sheet.cell(row=i+3, column=10).value >= 25:
                cnt7 += 1
                sheet.cell(row=i+3, column=7).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
            elif sheet.cell(row=i+3, column=10).value < 25 and sheet.cell(row=i+3, column=10).value >= 10:
                cnt8 += 1
                sheet.cell(row=i+3, column=7).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
            elif sheet.cell(row=i+3, column=10).value < 10 and sheet.cell(row=i+3, column=10).value >= 5:
                cnt9 += 1
                sheet.cell(row=i+3, column=7).fill = PatternFill(fgColor='00FF00', fill_type = "solid")
            elif sheet.cell(row=i+3, column=10).value < 5:
                cnt10 += 1
                sheet.cell(row=i+3, column=7).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")
                    

        

        ## Bordering Facility Level summary Sheet
        for i in range(2, len(dataframeForSheet6)+3):
            sheet.cell(row=i, column=2).border = thick_border
            sheet.cell(row=i, column=3).border = thick_border
            sheet.cell(row=i, column=4).border = thick_border
            sheet.cell(row=i, column=5).border = thick_border
            sheet.cell(row=i, column=6).border = thick_border
            sheet.cell(row=i, column=7).border = thick_border
            sheet.cell(row=i, column=8).border = thick_border


        # Coloring and palettes of Facility Guidance Sheet
        sheet['L32'] = 'Color Brackets'
        sheet['L33'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet['L34'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet['L35'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet['L36'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        sheet['L37'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet["M32"] = "Range"
        sheet["M33"] = ">= 50%"
        sheet["M34"] = "25 - 50%"
        sheet["M35"] = "10 - 25%"
        sheet["M36"] = "5 - 10%"
        sheet["M37"] = "< 5%"
        sheet["M38"] = "Total Facilities"

        sheet["N32"] = "Inconsistent"
        sheet["N33"] = cnt1
        sheet["N34"] = cnt2
        sheet["N35"] = cnt3
        sheet["N36"] = cnt4
        sheet["N37"] = cnt5
        sheet["N38"] = cnt1 + cnt2 + cnt3 + cnt4 + cnt5

        sheet["O32"] = "Inconsistent due to blank"
        sheet["O33"] = cnt6
        sheet["O34"] = cnt7
        sheet["O35"] = cnt8
        sheet["O36"] = cnt9
        sheet["O37"] = cnt10
        sheet["O38"] = cnt6 + cnt7 + cnt8 + cnt9 + cnt10


        ''' 
        GRAPH PLOTS
        '''
        
        Ranges = list(count_df_Test_Plot_Inc_State.keys())
        # Ranges = list(count_df_Test_Plot_INCBlnk_State.keys())
        Numbers_Inc = list(count_df_Test_Plot_Inc_State.values())
        Numbers_INCBlnk = list(count_df_Test_Plot_INCBlnk_State.values())


        #########################!-----------------------------------#############################!
        #########################!              PLOTS                #############################!
        #########################!-----------------------------------#############################!
        # TEST PLOT OF PIE CHART
        figure, axis = plt.subplots(2, 2)
        fig = plt.figure(figsize= (26, 8))
        rows = 2
        columns = 3


        #! 1 ....................................................................................
        fig.add_subplot(rows, columns, 1)
        X_axis = np.arange(len(Ranges))
        plt.bar(X_axis - 0.2, Numbers_Inc, 0.4, label = 'Number of Inconsistents')
        plt.bar(X_axis + 0.2, Numbers_INCBlnk, 0.4, label = 'Number of Inconsistent due to blank')
        #figure(figsize=(8, 8), dpi=50)
        plt.xticks(X_axis, Ranges)
        plt.xlabel("States")
        plt.ylabel("Number of Inconsistents and Inconsistent due to blank")
        plt.legend()
        plt.title('State Wise Error Summary')
        # set the spacing between subplots
        plt.subplots_adjust(left=0.1,
                            bottom=0.1, 
                            right=0.9, 
                            top=0.9, 
                            wspace=0.4, 
                            hspace=0.4)


        #! 2 ....................................................................................
        n = 2
        
        ###! GROUP BY STATE FOR 5 TOP NON-PERFORMER STATES  ---- & THEN 
        ###! GROUP BY DISTRICT FOR 5 TOP NON-PERFORMER DISTRICTS
        for _, g in final_result_summ2_For_SubPlots.groupby(['State']): 

            top_5_dist_INC = g.groupby(['District'])['Inconsistent'].sum().nlargest(5)
            top_5_dist_INCBlnk = g.groupby(['District'])['Inconsistent due to blank'].sum().nlargest(5)

            dist_dictionary_INC = top_5_dist_INC.to_dict()
            dist_dictionary_INCBlnk = top_5_dist_INCBlnk.to_dict()

            Ranges = list(dist_dictionary_INC.keys())
            Numbers_Inc = list(dist_dictionary_INC.values())
            Numbers_INCBlnk = list(dist_dictionary_INCBlnk.values())


            rows = 2
            columns = 3

            fig.add_subplot(rows, columns, n)
            
            X_axis = np.arange(len(Ranges))
            plt.bar(X_axis - 0.2, Numbers_INCBlnk, 0.4, label = 'Number of Inconsistent due to blank')
            plt.bar(X_axis - 0.2, Numbers_Inc, 0.4, label = 'Number of Inconsistents')

            #figure(figsize=(8, 8), dpi=50)
            plt.xticks(X_axis, Ranges)
            plt.xlabel(_)
            plt.ylabel("Number of Inconsistents and INCBlnk")
            plt.legend()
            plt.title('Top 5 non performing Districts')
            # set the spacing between subplots
            plt.subplots_adjust(left=0.1,
                                bottom=0.1, 
                                right=0.9, 
                                top=0.9, 
                                wspace=0.4, 
                                hspace=0.4)
   

            print('*********************** ' + str(n) + ' ######################################## ')

            n += 1

            if n > 6:
                break

        
        with tempfile.NamedTemporaryFile() as temp:
            plt.savefig(temp.name + ".png", dpi = 80)
            img = openpyxl.drawing.image.Image(temp.name + '.png')
            img.anchor='J1'
            sheet.add_image(img)
            temp.close()

        # ### delete useless columns of Facility Level Summary
        sheet.delete_cols(9, 2)

        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30

        workbook.save(filename=filename)

        '''
        **************************************************************************************************
        '''

        '''
        *********************************** sheet_4 :: Facility-wise Inconsistencies ***********************************
        '''

        # Checks Sheet (Inconsistent)
        # =======================================
        # workbook.active = sheet_4
        sheet_4['A1'] = "Facility-wise inconsistencies"
        sheet_4.oddHeader.center.size = 18
        sheet_4.oddHeader.center.font = "Tahoma,Bold"
        sheet_4.oddHeader.center.color = "CC3366"
        sheet_4['A1'].font = Font(size = 18, bold = True, color="003679")

        sheet_4.sheet_view.showGridLines = True
        for i in range(101):
            for j in range(len(FList3[i])):
                sheet_4.cell(row=j+8,column=i+2).value = FList3[i][j]
                
            # Alignment
            for k in range(1, len(FList3[i])+5):
                sheet_4.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')


        sheet_4['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_4['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_4['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_4['A2'].font = Font(color = "FFFFFF")
        sheet_4['A3'].font = Font(color = "FFFFFF")
        sheet_4['A4'].font = Font(color = "FFFFFF")
        # set the width of the row
        sheet_4.column_dimensions['A'].width = 45
        sheet_4.row_dimensions[2].height = 20
        sheet_4.row_dimensions[3].height = 20
        sheet_4.row_dimensions[4].height = 20

        # ## Bordering Checks Giving Inconsistent Sheet
        # for i in range(1, len(dataframeForSheet6)+50):
        #     sheet_4.cell(row=1, column=i).border = thick_border
        #     sheet_4.cell(row=2, column=i).border = thick_border
        #     sheet_4.cell(row=1, column=i).border = thick_border

        workbook.save(filename=filename)

        '''
        **************************************************************************************************
        '''

        '''
        *********************************** sheet_5 :: INCBlnk-wise Inconsistencies ***********************************
        '''

        # Checks Sheet (INCBlnk)
        # =======================================

        sheet_5['A1'] = "Facility-wise Inconsistencies due to blank"
        sheet_5.oddHeader.center.size = 18
        sheet_5.oddHeader.center.font = "Tahoma,Bold"
        sheet_5.oddHeader.center.color = "CC3366"
        sheet_5['A1'].font = Font(size = 18, bold = True, color="003679")
        sheet_5.column_dimensions['A'].width = 55

        sheet_5.sheet_view.showGridLines = True
        for i in range(101):
            for j in range(len(FList4_copy[i])):
                sheet_5.cell(row=j+8,column=i+2).value = FList4_copy[i][j]
                # print(sheet_5.cell(row=j+8,column=i+2).value)

            # Alignment
            for k in range(1, len(FList4[i])+5):
                sheet_5.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')

        sheet_5['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_5['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_5['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_5['A2'].font = Font(color = "FFFFFF")
        sheet_5['A3'].font = Font(color = "FFFFFF")
        sheet_5['A4'].font = Font(color = "FFFFFF")
        sheet_5.column_dimensions['A'].width = 25
        sheet_5.row_dimensions[2].height = 20
        sheet_5.row_dimensions[3].height = 20
        sheet_5.row_dimensions[4].height = 20 

        workbook.save(filename=filename)

        '''
        **************************************************************************************************
        '''

        '''
        *********************************** sheet_1 :: Validation rule wise summary ***********************************
        '''

        ### Formatting Sheet_1{Validation Checkwise Summary}
        try:
            sheet_1.move_range("A1:G10000", rows=1)
        except:
            sheet_1.move_range("A1:G25000", rows=1)
            try:
                sheet_1.move_range("A1:G50000", rows=1)
            except:
                sheet_1.move_range("A1:G100000", rows=1)
            finally:
                sheet_1.move_range("A1:G1048576", rows=1)


        sheet_1.oddHeader.center.text = "Validation Rule Wise Summary"
        sheet_1.cell(row=2, column=1).value = 'Sr. No'
        sheet_1.oddHeader.center.size = 14
        sheet_1.oddHeader.center.font = "Tahoma,Bold"
        sheet_1.oddHeader.center.color = "CC3366"
        sheet_1.cell(row=1, column=2).value = 'Validation Rule Wise Summary'
        sheet_1['B1'].font = Font(size = 18, bold = True, color="003679")

        ## EXPANDING ROWS OF SHEET_1 (VALIDATION CHECKWISE SUMMARY)
        sheet_1.column_dimensions['B'].width = 20
        sheet_1.column_dimensions['C'].width = 20
        sheet_1.column_dimensions['D'].width = 15
        sheet_1.column_dimensions['E'].width = 25

        # HYPERLINKING FUNCTIONS
        def excel_cols_link1():
            n1 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n1))
                n1 += 1

        # HYPERLINKING FUNCTIONS
        def excel_cols_link2():
            n2 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n2))
                n2 += 1
        
        cnt21, cnt22, cnt23, cnt24 = 0, 0, 0, 0
        for i in range(len(FList1)):
            if sheet_1.cell(row=i+3, column=6).value == None:
                sheet.delete_rows(i+3, 1)
            elif sheet_1.cell(row=i+3, column=6).value >= 25:
                cnt21 += 1
            elif sheet_1.cell(row=i+3, column=6).value < 25 and sheet_1.cell(row=i+3, column=6).value >= 10:
                cnt22 += 1
            elif sheet_1.cell(row=i+3, column=6).value < 10 and sheet_1.cell(row=i+3, column=6).value >= 5:
                cnt23 += 1
            elif sheet_1.cell(row=i+3, column=6).value < 5:
                cnt24 += 1

            for j in range(len(FList1[i])):

                # ### RESPONSIBLE FOR EXPANDING FACILITY NAMES IN THE  (Inconsitency due to blank-wise facilities)
                ##### FOR SHEET_2 #####
                sheet_2.cell(row=j+5,column=i+2).value = FList1[i][j]

                '''
                # Create hyperlink to relevant cell for sheet_1
                '''
                link1 = "#'Inconsistency-wise facilities'!B2"

                l1 = list(itertools.islice(excel_cols_link1(), len(FList1)+1))
                
                ## update link and then hyperlink it
                link1 = link1.replace("B", l1[i+1])
                sheet_1.cell(row=i+3, column=4).hyperlink = link1
                sheet_1.cell(row=i+3, column=4).style = "Hyperlink"

                if sheet_1.cell(row=i+3, column=6).value >= 25:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=6).value < 25 and sheet_1.cell(row=i+3, column=6).value >= 10:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=6).value < 10 and sheet_1.cell(row=i+3, column=6).value >= 5:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=6).value < 5:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")


        link3 = "#'Inc due to blank-wise facility'!B2"
        l3 = list(itertools.islice(excel_cols_link2(), dataframeForSheet5.shape[0]+1))

        ##### Generating links 
        k = 0
        for i in range(len(FList2)):

            if sheet_1.cell(row=i+3, column=5).value == 0:
                k = k + 1 
            if sheet_1.cell(row=i+3, column=5).value != 0 and sheet_1.cell(row=i+2, column=5).value == 0:
                if k == 1:
                    link3 = "#'Inc due to blank-wise facility'!"+ l3[i] +"2"
                    print('i=', i, "cell=", sheet_1.cell(row=i+3, column=5), "Value=", sheet_1.cell(row=i+3, column=5).value, 'link=', link3)
                elif k > 1:
                    link3 = "#'Inc due to blank-wise facility'!"+ l3[i+1-k] +"2"
                    print('i=', i, "cell=", sheet_1.cell(row=i+3, column=5), "Value=", sheet_1.cell(row=i+3, column=5).value, 'link=', link3)

            elif sheet_1.cell(row=i+3, column=5).value != 0 and sheet_1.cell(row=i+2, column=5).value != 0:
                link3 = "#'Inc due to blank-wise facility'!"+ l3[i+1-k] +"2"
                print('i=', i, "celltype2=", sheet_1.cell(row=i+3, column=5), "Value=", sheet_1.cell(row=i+3, column=5).value, 'link=', link3)
            else:
                link3=None

            sheet_1.cell(row=i+3, column=5).hyperlink = link3
            sheet_1.cell(row=i+3, column=5).style = "Hyperlink"


        cnt25, cnt26, cnt27, cnt28 = 0, 0, 0, 0
        for i in range(len(FList2)):
            if sheet_1.cell(row=i+3, column=7).value == None:
                sheet_1.delete_rows(i+3, 1)
            elif sheet_1.cell(row=i+3, column=7).value >= 25:
                cnt25 += 1
                sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
            elif sheet_1.cell(row=i+3, column=7).value < 25 and sheet_1.cell(row=i+3, column=7).value >= 10:
                cnt26 += 1
                sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
            elif sheet_1.cell(row=i+3, column=7).value < 10 and sheet_1.cell(row=i+3, column=7).value >= 5:
                cnt27 += 1
                sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
            elif sheet_1.cell(row=i+3, column=7).value < 5:
                cnt28 += 1
                sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")

        ### RESPONSIBLE FOR EXPANDING FACILITY NAMES IN THE  (Inconsitency due to blank-wise facilities)
        for i in range(dataframeForSheet5.shape[0]):   
            ########### FILLING VALUES #####################
            for j in range(len(FList2_copy[i])):
                sheet_3.cell(row=j+5,column=i+2).value = FList2_copy[i][j]

        
        ### Coloring of Validation Summary Sheet
        sheet_1['J5'] = "Color Brackets"
        sheet_1['J6'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet_1['J7'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet_1['J8'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet_1['J9'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet_1['K5'] = "Range"
        sheet_1['K6'] = ">= 25%"
        sheet_1['K7'] = "10 - 25%"
        sheet_1['K8'] = "5 - 10%"
        sheet_1['K9'] = "< 5%"
        sheet_1['K10'] = "Total Indicators"

        sheet_1['L5'] = "Inconsistent"
        sheet_1['L6'] = cnt21
        sheet_1['L7'] = cnt22
        sheet_1['L8'] = cnt23
        sheet_1['L9'] = cnt24
        sheet_1['L10'] = cnt21 + cnt22 + cnt23 + cnt24

        sheet_1['M5'] = "Inconsistent due to blank"
        sheet_1['M6'] = cnt25
        sheet_1['M7'] = cnt26
        sheet_1['M8'] = cnt27
        sheet_1['M9'] = cnt28
        sheet_1['M10'] = cnt25 + cnt26 + cnt27 + cnt28


        ## Bordering Validation checkwise summary Sheet
        for i in range(2, len(dataframeForSheet4)+3):
            sheet_1.cell(row=i, column=2).border = thick_border
            sheet_1.cell(row=i, column=3).border = thick_border
            sheet_1.cell(row=i, column=4).border = thick_border
            sheet_1.cell(row=i, column=5).border = thick_border

        ### delete useless columns of sheet_1
        sheet_1.delete_cols(6, 2)  

        workbook.save(filename=filename)

        '''
        **************************************************************************************************
        '''

        '''
        *********************************** sheet_2 :: Inconsistency-wise facilities ***********************************
        '''

        # # Attention Required Sheet (Inconsistent)
        # # =======================================
        sheet_2.sheet_view.showGridLines = True

        sheet_2['A1'] = "Inconsistency-wise facilities"
        sheet_2.oddHeader.center.size = 18
        sheet_2.oddHeader.center.font = "Tahoma,Bold"
        sheet_2.oddHeader.center.color = "CC3366"
        sheet_2['A1'].font = Font(size = 18, bold = True, color="003679")

        ## RESPONSIBLE FOR EXPANDING FACILITY NAMES IN THE  (Inconsistency-wise facilities)
        for j in range(len(FList1[i])):
            sheet_2.cell(row=j+5,column=i+2).value = FList1[i][j]

        sheet_2['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_2['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_2['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_2['A5'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_2['A2'].font = Font(color = "FFFFFF")
        sheet_2['A3'].font = Font(color = "FFFFFF")
        sheet_2['A4'].font = Font(color = "FFFFFF")
        sheet_2['A5'].font = Font(color = "FFFFFF")
        sheet_2.column_dimensions['A'].width = 35
        sheet_2.row_dimensions[2].height = 20
        sheet_2.row_dimensions[3].height = 20
        sheet_2.row_dimensions[4].height = 20
        sheet_2.row_dimensions[5].height = 20   
                
        workbook.save(filename=filename)

        '''
        **************************************************************************************************
        '''

        '''
        *********************************** sheet_3 :: Inconsitency due to blank-wise facilities ***********************************
        '''

        sheet_3.sheet_view.showGridLines = True
        sheet_3['A1'] = "Inconsitency due to blank-wise facilities"
        sheet_3.oddHeader.center.size = 18
        sheet_3.oddHeader.center.font = "Tahoma,Bold"
        sheet_3.oddHeader.center.color = "CC3366"
        sheet_3['A1'].font = Font(size = 18, bold = True, color="003679")

        ### RESPONSIBLE FOR EXPANDING FACILITY NAMES IN THE  (Inconsitency due to blank-wise facilities)
        for i in range(dataframeForSheet5.shape[0]):   
            ########### FILLING VALUES #####################
            for j in range(len(FList2_copy[i])):
                sheet_3.cell(row=j+5,column=i+2).value = FList2_copy[i][j]
                
            # Colors
            for k in range(1, len(FList2[i])+100):
                sheet_3.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                # sheet_3.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")

        sheet_3['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A5'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A2'].font = Font(color = "FFFFFF")
        sheet_3['A3'].font = Font(color = "FFFFFF")
        sheet_3['A4'].font = Font(color = "FFFFFF")
        sheet_3['A5'].font = Font(color = "FFFFFF")
        sheet_3.column_dimensions['A'].width = 35
        sheet_3.row_dimensions[2].height = 20
        sheet_3.row_dimensions[3].height = 20
        sheet_3.row_dimensions[4].height = 20
        sheet_3.row_dimensions[5].height = 20 
        
        print('@@@@@ Code is reaching till the last sheet @@@@@@@')
        print('@@@@@ Code is reaching till the last sheet @@@@@@@')
        print('@@@@@ Code is reaching till the last sheet @@@@@@@')

        workbook.save(filename=filename)
        

        '''
        **************************************************************************************************
        '''

        from PyQt5.QtCore import QTimer

        # Using QTimer to delay the QMessageBox
        QTimer.singleShot(1000, self.show_message_box)

        print("!!!!!!!!!!!!!!!!!!! $$$$$$$$$$$$$$$$$$$$ ******************** $$$$$$$$$$$$$$$$$$$$ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")

        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon(self.resource_path('checked.png')))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("Excel file downloaded in the selected location \n\n एक्सेल फ़ाइल चयनित स्थान पर डाउनलोड की गई।")
        # Set the title of the window
        self.msg.setWindowTitle(" Successful Download ")
        # Display the message box
        self.msg.show()

        # os.remove("FileName.csv")
        # os.remove("myplot2.png")


    # Reset
    def reset(self):
        QtCore.QCoreApplication.quit()
        status = QtCore.QProcess.startDetached(sys.executable, sys.argv)
      
    # Display methodology pdf in browser
    def UserManualEnglish(self):
        os.system('start Manual_ADVTool_English.pdf')

    def UserManualHindi(self):
        os.system('start Manual_ADVTool_English.pdf')


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    
    TabWidget = QtWidgets.QTabWidget()
    ui = Ui_TabWidget()
    ui.setupUi(TabWidget)
    TabWidget.show()
    sys.exit(app.exec_())
